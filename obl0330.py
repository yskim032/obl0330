import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import json
from typing import List, Dict
from tkinterdnd2 import DND_FILES, TkinterDnD
from datetime import datetime
from openpyxl import load_workbook
import xlrd  # .xls 파일 처리를 위한 라이브러리
import time
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl import utils
import re

# pyinstaller -w -F --add-binary="C:/Users/kod03/AppData/Local/Programs/Python/Python311/tcl/tkdnd2.8;tkdnd2.8" obl0330.py

class ContainerConverter:
    # 전역 변수 선언
    OH = "0"
    Oleft = "0"
    Oright = "0"
    
    def __init__(self):
        self.root = TkinterDnD.Tk()
        self.root.title("CLL to OBL Converter")
        self.root.geometry("1000x900")

        # 설정 파일 경로 설정
        self.desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        self.onedrive_desktop = os.path.join(os.path.expanduser('~'), 'OneDrive', '바탕 화면')
        self.config_dir = os.path.join(self.desktop_path, "OBL_Configs")
        
        # port_codes 딕셔너리 추가
        self.port_codes = {'AEAJM': 'AJMAN', 'AEAUH': 'ABU DHABI', 'AEDXB': 'DUBAI', 'AEFJR': 'AL - FUJAYRAH', 'AEJEA': 'JEBEL ALI', 'AEKLF': 'KHOR AL FAKKAN', 'AEPRA': 'PORT RASHID', 'AEQIW': 'UMM AL QAIWAIN', 'AERKT': 'RAS AL KHAIMAH', 'AESHJ': 'SHARJAH', 'AEYAS': 'YAS ISLAND', 'AGANU': 'ANTIGUA', 'AIRBY': 'ROAD BAY', 'ALDRZ': 'DURRES', 'ALSAR': 'SARANDE', 'AOLAD': 'LUANDA', 'AOLOB': 'LOBITO', 'AOMSZ': 'NAMIBE', 'ARBHI': 'BAHIA BLANCA', 'ARBUE': 'BUENOS AIRES', 'ARCMP': 'CAMPANA', 'ARCNQ': 'CORRIENTES', 'ARLPG': 'LA PLATA', 'ARMDQ': 'MAR DEL PLATA', 'ARPMY': 'PUERTO MADRYN', 'ARPSS': 'POSADAS', 'ARROS': 'ROSARIO', 'ARSAE': 'SAN ANTONIO ESTE', 'ARUSH': 'USHUAIA', 'ARZAE': 'ZARATE', 'ASPPG': 'PAGO PAGO', 'AUABP': 'ABBOT POINT', 'AUADL': 'ADELAIDE', 'AUALH': 'ALBANY', 'AUBEL': 'BELL BAY', 'AUBNE': 'BRISBANE', 'AUBOO': 'BOOBY ISLAND', 'AUCNS': 'CAIRNS', 'AUDRW': 'DARWIN', 'AUEPR': 'ESPERANCE', 'AUFRE': 'FREMANTLE', 'AUGLT': 'GLADSTONE', 'AUHBA': 'HOBART', 'AUHPT': 'HAY POINT', 'AUMEL': 'MELBOURNE', 'AUNTL': 'NEWCASTLE', 'AUPHE': 'PORT HEDLAND', 'AUPKL': 'PORT KEMBLA', 'AUSYD': 'SYDNEY', 'AWORJ': 'ORANJESTAD', 'BBBGI': 'BRIDGETOWN', 'BDCGP': 'CHATTOGRAM', 'BDMGL': 'MONGLA', 'BEANR': 'ANTWERP', 'BEGNE': 'GENT (GHENT)', 'BEZEE': 'ZEEBRUGGE', 'BGBOJ': 'BURGAS', 'BGVAR': 'VARNA', 'BHKBS': 'BAHRAIN', 'BJCOO': 'COTONOU', 'BMBDA': 'HAMILTON', 'BMKWF': 'KINGS WHARF', 'BNMUA': 'MUARA', 'BQEUX': 'SINT EUSTATIUS', 'BRACB': 'ARRAIAL DO CABO', 'BRADR': 'ANGRA DOS REIS', 'BRANG': 'ARTUR NOGUEIRA', 'BRBEL': 'BELEM', 'BRBZC': 'BUZIOS', 'BRCBU': 'CAMBORIU', 'BRCDO': 'CABEDELO', 'BRCOP': 'CARMO DO PARANAIBA', 'BRFOR': 'FORTALEZA', 'BRIBB': 'IMBITUBA', 'BRIBE': 'ILHABELA', 'BRIGE': 'ILHA GRANDE', 'BRIGI': 'ITAGUAI', 'BRIOA': 'ITAPOA', 'BRIOS': 'ILHEUS', 'BRIQI': 'ITAQUI', 'BRITA': 'ITACOATIARA', 'BRITJ': 'ITAJAI', 'BRMAO': 'MANAUS', 'BRMCZ': 'MACEIO', 'BRNVT': 'NAVEGANTES', 'BRPBO': 'PORTO BELO', 'BRPEC': 'PECEM', 'BRPNG': 'PARANAGUA', 'BRPVH': 'PORTO VELHO', 'BRQCK': 'CABO FRIO', 'BRREC': 'RECIFE', 'BRRIG': 'RIO GRANDE', 'BRRIO': 'RIO DE JANEIRO', 'BRSFS': 'SAO FRANCISCO DO SUL', 'BRSSA': 'SALVADOR', 'BRSSZ': 'SANTOS', 'BRSTM': 'SANTAREM', 'BRSUA': 'SUAPE', 'BRUBT': 'UBATUBA', 'BRVIX': 'VITORIA', 'BRVLC': 'VILA DO CONDE', 'BSCOC': 'COCO CAY'}
        
        # 설정 파일 찾기 및 로드
        self.find_and_load_config_files()
        
        # POL, TOL 선택 값 저장 변수
        self.selected_pol = tk.StringVar()
        self.selected_tol = tk.StringVar()

        # ITPS 관련 변수
        self.itps_file = None
        self.obl_file = None

        self.setup_ui()
        self.reset_all()

    def find_and_load_config_files(self):
        """설정 파일 찾기 및 로드"""
        # 설정 파일 이름
        stow_filename = "StowCodes_mapping.json"
        tpsz_filename = "SZTP_mapping.json"
        
        # 파일 찾기
        self.stowage_config_file = self.find_config_file(stow_filename)
        self.tpsz_config_file = self.find_config_file(tpsz_filename)
        
        print(f"Stowage config file path: {self.stowage_config_file}")  # 디버깅용
        
        # 설정 로드
        self.stowage_settings = self.load_stowage_settings()
        # print(f"Loaded stowage settings: {self.stowage_settings}")  # 디버깅용
        
        # stow_mapping을 직접 설정값으로 설정 (중간 딕셔너리 없이)
        self.stow_mapping = self.stowage_settings
        # print(f"Stow mapping: {self.stow_mapping}")  # 디버깅용
        
        # TPSZ 설정 로드
        self.tpsz_settings = self.load_tpsz_settings()
        self.tpsz_mapping = self.tpsz_settings.get('mapping', {})
        self.tpsz_column_mapping = self.tpsz_settings.get('column_mapping', {
            'before': 'Description',
            'after': 'Code'
        })

        # 서비스 목록 업데이트를 위해 변수 준비
        self.selected_service = tk.StringVar()

    def find_config_file(self, filename: str) -> str:
        """설정 파일 찾기"""
        # 가능한 경로들
        possible_paths = [
            os.path.join(self.desktop_path, filename),  # 일반 바탕화면
            os.path.join(self.onedrive_desktop, filename),  # OneDrive 바탕화면
            os.path.join(self.config_dir, filename)  # 설정 디렉토리
        ]
        
        # 존재하는 파일 찾기
        for path in possible_paths:
            if os.path.exists(path):
                print(f"Found config file: {path}")  # 디버깅용
                return path
        
        # 파일이 없으면 설정 디렉토리에 생성
        os.makedirs(self.config_dir, exist_ok=True)
        default_path = os.path.join(self.config_dir, filename)
        with open(default_path, 'w', encoding='utf-8') as f:
            json.dump({}, f, ensure_ascii=False, indent=2)
        print(f"Created new config file: {default_path}")  # 디버깅용
        
        # 사용자에게 알림
        messagebox.showinfo(
            "설정 파일 생성",
            f"설정 파일을 찾을 수 없어 새로 생성했습니다:\n{default_path}\n"
            f"바탕화면에 {filename} 파일이 있다면 {self.config_dir} 폴더로 복사해주세요."
        )
        
        return default_path

    def load_stowage_settings(self) -> Dict:
        """Stowage 설정 로드"""
        try:
            with open(self.stowage_config_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # print("Loaded Stowage settings:", data)  # 디버깅용
                return data
        except Exception as e:
            print(f"Error loading Stowage settings: {str(e)}")  # 디버깅용
            messagebox.showerror("Error", f"Stowage 매핑 파일 로드 실패: {str(e)}")
            return {}

    def load_tpsz_settings(self) -> Dict:
        """TpSz 설정 로드"""
        try:
            with open(self.tpsz_config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            messagebox.showerror("Error", f"TpSz 매핑 파일 로드 실패: {str(e)}")
            return {}

    def setup_ui(self):
        """UI 설정"""
        # 탭 컨트롤 생성
        self.tab_control = ttk.Notebook(self.root)
        self.tab_control.pack(expand=True, fill="both")
        
        # 각 탭 프레임 생성
        self.single_tab = ttk.Frame(self.tab_control)
        self.multi_cll_tab = ttk.Frame(self.tab_control)
        self.itps_tab = ttk.Frame(self.tab_control)
        self.stowage_tab = ttk.Frame(self.tab_control)
        self.tpsz_tab = ttk.Frame(self.tab_control)
        self.edi_tab = ttk.Frame(self.tab_control)  # EDI PARSER 탭 추가
        
        # 탭 추가
        self.tab_control.add(self.single_tab, text='단일 CLL 변환')
        self.tab_control.add(self.multi_cll_tab, text='Multi CLL 변환')
        self.tab_control.add(self.itps_tab, text='ITPS 추가')
        self.tab_control.add(self.stowage_tab, text='STOWAGE CODE 관리')
        self.tab_control.add(self.tpsz_tab, text='TpSZ 관리')
        self.tab_control.add(self.edi_tab, text='EDI PARSER')  # EDI PARSER 탭 추가
        
        # 각 탭 설정
        self.setup_single_tab()
        self.setup_multi_cll_tab()
        self.setup_itps_tab()
        self.setup_stowage_tab()
        self.setup_tpsz_tab()
        self.setup_edi_tab()  # EDI PARSER 탭 설정 메서드 호출
        
        # JSON 파일 내용 표시
        self.update_stowage_preview()  # Stowage 탭 업데이트
        self.update_tpsz_preview()     # TpSZ 탭 업데이트

    def setup_single_tab(self):
        # 단일 CLL 변환 탭 설정
        left_frame = ttk.Frame(self.single_tab)
        left_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        right_frame = ttk.Frame(self.single_tab)
        right_frame.pack(side="right", fill="both", padx=5)
        
        # POL, TOL 선택 프레임
        port_frame = ttk.LabelFrame(left_frame, text="POL TOL")
        port_frame.pack(pady=10, padx=10, fill="x")

        # POL 버튼 프레임
        pol_frame = ttk.LabelFrame(port_frame, text="POL")
        pol_frame.pack(pady=5, padx=5, fill="x")

        # POL 버튼들
        pol_ports = ['KRPUS', 'KRKAN', 'KRINC']
        self.pol_buttons = {}
        for port in pol_ports:
            btn = tk.Button(pol_frame, text=port, width=10,
                          command=lambda p=port: self.select_pol(p))
            btn.pack(side=tk.LEFT, padx=5, pady=5)
            self.pol_buttons[port] = btn

        # TOL 버튼 프레임
        tol_frame = ttk.LabelFrame(port_frame, text="TOL")
        tol_frame.pack(pady=5, padx=5, fill="x")

        # TOL 버튼들과 매핑
        tol_mapping = {
            'PNC': 'KRPUSPN',
            'PNIT': 'KRPUSAB',
            'BCT': 'KRPUSBC',
            'HJNC': 'KRPUSAP',
            'ICT': 'KRINCAH',
            'GWCT': 'KRKANKT'
        }
        self.tol_buttons = {}
        self.tol_values = tol_mapping
        for btn_text, value in tol_mapping.items():
            btn = tk.Button(tol_frame, text=btn_text, width=10,
                          command=lambda v=value: self.select_tol(v))
            btn.pack(side=tk.LEFT, padx=5, pady=5)
            self.tol_buttons[value] = btn

        # 파일 정보 표시 영역
        info_frame = ttk.LabelFrame(left_frame, text="파일 정보")
        info_frame.pack(pady=10, padx=10, fill="x")

        self.input_label = ttk.Label(info_frame, text="입력 파일: 없음")
        self.input_label.pack(pady=5, anchor="w")

        self.output_label = ttk.Label(info_frame, text="출력 파일: 없음")
        self.output_label.pack(pady=5, anchor="w")

        # CLL 변환을 위한 드래그 & 드롭 영역
        self.cll_frame = ttk.LabelFrame(left_frame, text="CLL -> OBL 변환")
        self.cll_frame.pack(pady=10, padx=10, fill="x")

        self.cll_label = ttk.Label(self.cll_frame, text="CLL 파일을 여기에 드롭하세요")
        self.cll_label.pack(pady=20)

        # CLL 드래그 앤 드롭 바인딩
        self.cll_label.drop_target_register(DND_FILES)
        self.cll_label.dnd_bind('<<Drop>>', self.drop_cll_file)

        # OBL EMPTY 추가를 위한 드래그 & 드롭 영역
        self.obl_frame = ttk.LabelFrame(left_frame, text="OBL EMPTY 추가")
        self.obl_frame.pack(pady=10, padx=10, fill="x")

        self.obl_label = ttk.Label(self.obl_frame, text="OBL 파일을 여기에 드롭하세요")
        self.obl_label.pack(pady=20)

        # OBL 드래그 앤 드롭 바인딩
        self.obl_label.drop_target_register(DND_FILES)
        self.obl_label.dnd_bind('<<Drop>>', self.drop_obl_file)

        # EMPTY 컨테이너 입력 섹션
        empty_frame = ttk.LabelFrame(left_frame, text="EMPTY 컨테이너 추가")
        empty_frame.pack(pady=10, padx=10, fill="x")

        # 5개의 입력 행 생성
        self.empty_entries = []
        for i in range(5):
            row_frame = ttk.Frame(empty_frame)
            row_frame.pack(pady=5)

            pod_entry = ttk.Entry(row_frame, width=10)
            pod_entry.pack(side="left", padx=5)
            pod_entry.insert(0, "POD")
            pod_entry.bind('<FocusIn>', lambda e, entry=pod_entry: self.on_entry_click(e, entry))
            pod_entry.bind('<FocusOut>', lambda e, entry=pod_entry: self.on_focus_out(e, entry, "POD"))
            pod_entry.bind('<Key>', lambda e, entry=pod_entry: self.on_key_press(e, entry))

            sztp_entry = ttk.Entry(row_frame, width=10)
            sztp_entry.pack(side="left", padx=5)
            sztp_entry.insert(0, "SzTp")
            sztp_entry.bind('<FocusIn>', lambda e, entry=sztp_entry: self.on_entry_click(e, entry))
            sztp_entry.bind('<FocusOut>', lambda e, entry=sztp_entry: self.on_focus_out(e, entry, "SzTp"))
            sztp_entry.bind('<Key>', lambda e, entry=sztp_entry: self.on_key_press(e, entry))

            qty_entry = ttk.Entry(row_frame, width=5)
            qty_entry.pack(side="left", padx=5)
            qty_entry.insert(0, "수량")
            qty_entry.bind('<FocusIn>', lambda e, entry=qty_entry: self.on_entry_click(e, entry))
            qty_entry.bind('<FocusOut>', lambda e, entry=qty_entry: self.on_focus_out(e, entry, "수량"))
            qty_entry.bind('<Key>', lambda e, entry=qty_entry: self.on_key_press(e, entry))

            self.empty_entries.append((pod_entry, sztp_entry, qty_entry))

        # Summary 표시 영역을 right_frame으로 이동
        self.single_summary_frame = ttk.LabelFrame(right_frame, text="Container Summary")
        self.single_summary_frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        self.single_summary_text = tk.Text(self.single_summary_frame, height=30, width=40)
        self.single_summary_text.pack(pady=5, padx=5, fill="both", expand=True)
        self.single_summary_text.insert(tk.END, "단일 CLL 탭에서 파일 변환 시 Summary가 표시됩니다.")

    def setup_multi_cll_tab(self):
        """CLL 파일 병합 탭 설정"""
        # 좌우 분할
        left_frame = ttk.Frame(self.multi_cll_tab)
        left_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        right_frame = ttk.Frame(self.multi_cll_tab)
        right_frame.pack(side="right", fill="both", padx=5)
        
        # POL/TOL 선택 프레임
        port_frame = ttk.LabelFrame(left_frame, text="POL TOL")
        port_frame.pack(pady=10, padx=10, fill="x")
        
        # POL 버튼 프레임
        pol_frame = ttk.LabelFrame(port_frame, text="POL")
        pol_frame.pack(pady=5, padx=5, fill="x")
        
        pol_ports = ['KRPUS', 'KRKAN', 'KRINC']
        self.multi_pol_buttons = {}
        for port in pol_ports:
            btn = tk.Button(pol_frame, text=port, width=10,
                          command=lambda p=port: self.select_multi_pol(p))
            btn.pack(side=tk.LEFT, padx=5, pady=5)
            self.multi_pol_buttons[port] = btn

        # TOL 버튼 프레임
        tol_frame = ttk.LabelFrame(port_frame, text="TOL")
        tol_frame.pack(pady=5, padx=5, fill="x")
        
        tol_mapping = {
            'PNC': 'KRPUSPN',
            'PNIT': 'KRPUSAB',
            'BCT': 'KRPUSBC',
            'HJNC': 'KRPUSAP',
            'ICT': 'KRINCAH',
            'GWCT': 'KRKANKT'
        }
        
        self.multi_tol_buttons = {}
        for btn_text, value in tol_mapping.items():
            btn = tk.Button(tol_frame, text=btn_text, width=10,
                          command=lambda v=value: self.select_multi_tol(v))
            btn.pack(side=tk.LEFT, padx=5, pady=5)
            self.multi_tol_buttons[btn_text] = btn

        # 파일 선택 영역 컨테이너
        files_frame = ttk.Frame(left_frame)
        files_frame.pack(pady=10, padx=10, fill="x")

        # Master CLL 파일 프레임
        self.master_frame = ttk.LabelFrame(files_frame, text="첫 번째(Master) CLL 파일")
        self.master_frame.pack(pady=5, padx=5, fill="x")
        
        self.master_label = ttk.Label(self.master_frame, text="CLL 파일을 여기에 드롭하세요")
        self.master_label.pack(pady=10)
        
        self.master_path_label = ttk.Label(self.master_frame, text="파일 경로: 없음")
        self.master_path_label.pack(pady=5)
        
        # Master 파일 드롭 영역 바인딩
        self.master_frame.drop_target_register(DND_FILES)
        self.master_frame.dnd_bind('<<Drop>>', self.drop_master_cll)

        # Slave CLL 파일 프레임
        self.slave_frame = ttk.LabelFrame(files_frame, text="두 번째(Slave) CLL 파일")
        self.slave_frame.pack(pady=5, padx=5, fill="x")
        
        self.slave_label = ttk.Label(self.slave_frame, text="CLL 파일을 여기에 드롭하세요")
        self.slave_label.pack(pady=10)
        
        self.slave_path_label = ttk.Label(self.slave_frame, text="파일 경로: 없음")
        self.slave_path_label.pack(pady=5)
        
        # Slave 파일 드롭 영역 바인딩
        self.slave_frame.drop_target_register(DND_FILES)
        self.slave_frame.dnd_bind('<<Drop>>', self.drop_slave_cll)

        # 결과 정보 프레임
        self.result_frame = ttk.LabelFrame(right_frame, text="변환 결과")
        self.result_frame.pack(pady=10, padx=10, fill="x")
        
        self.result_label = ttk.Label(self.result_frame, text="출력 파일: 없음")
        self.result_label.pack(pady=5)

        # Summary 표시 영역을 right_frame으로 이동
        self.multi_summary_frame = ttk.LabelFrame(right_frame, text="Container Summary")
        self.multi_summary_frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        self.multi_summary_text = tk.Text(self.multi_summary_frame, height=30, width=40)
        self.multi_summary_text.pack(pady=5, padx=5, fill="both", expand=True)
        self.multi_summary_text.insert(tk.END, "Multi CLL 탭에서 파일 변환 시 Summary가 표시됩니다.")

    def setup_itps_tab(self):
        """ITPS 추가 탭 설정"""
        # 좌우 분할
        left_frame = ttk.Frame(self.itps_tab)
        left_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        right_frame = ttk.Frame(self.itps_tab)
        right_frame.pack(side="right", fill="both", padx=5)

        # 파일 정보 표시 영역
        info_frame = ttk.LabelFrame(left_frame, text="파일 정보")
        info_frame.pack(pady=10, padx=10, fill="x")

        self.itps_input_label = ttk.Label(info_frame, text="ITPS 파일: 없음")
        self.itps_input_label.pack(pady=5, anchor="w")

        self.itps_obl_label = ttk.Label(info_frame, text="OBL 파일: 없음")
        self.itps_obl_label.pack(pady=5, anchor="w")

        self.itps_output_label = ttk.Label(info_frame, text="출력 파일: 없음")
        self.itps_output_label.pack(pady=5, anchor="w")

        # ITPS 파일 드롭 영역
        itps_drop_frame = ttk.LabelFrame(left_frame, text="ITPS 파일 드롭")
        itps_drop_frame.pack(pady=10, padx=10, fill="x")

        self.itps_drop_label = ttk.Label(itps_drop_frame, text="ITPS 파일을 여기에 드롭하세요")
        self.itps_drop_label.pack(pady=20)

        # ITPS 드래그 앤 드롭 바인딩
        self.itps_drop_label.drop_target_register(DND_FILES)
        self.itps_drop_label.dnd_bind('<<Drop>>', self.drop_itps_file)

        # OBL 파일 드롭 영역
        obl_drop_frame = ttk.LabelFrame(left_frame, text="OBL 파일 드롭")
        obl_drop_frame.pack(pady=10, padx=10, fill="x")

        self.obl_drop_label = ttk.Label(obl_drop_frame, text="OBL 파일을 여기에 드롭하세요")
        self.obl_drop_label.pack(pady=20)

        # OBL 드래그 앤 드롭 바인딩
        self.obl_drop_label.drop_target_register(DND_FILES)
        self.obl_drop_label.dnd_bind('<<Drop>>', self.drop_obl_for_itps)

        # Summary 표시 영역
        self.itps_summary_frame = ttk.LabelFrame(right_frame, text="ITPS Summary")
        self.itps_summary_frame.pack(pady=10, padx=10, fill="both", expand=True)
        
        self.itps_summary_text = tk.Text(self.itps_summary_frame, height=30, width=40)
        self.itps_summary_text.pack(pady=5, padx=5, fill="both", expand=True)
        self.itps_summary_text.insert(tk.END, "ITPS 파일 처리 시 Summary가 표시됩니다.")

    def setup_stowage_tab(self):
        """STOWAGE CODE 관리 탭 설정"""
        # 메인 프레임
        main_frame = ttk.Frame(self.stowage_tab)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 드래그 & 드롭 영역
        drop_frame = ttk.LabelFrame(main_frame, text="Stowage Code 엑셀 파일")
        drop_frame.pack(fill="x", pady=(0, 10))

        self.stowage_drop_label = ttk.Label(drop_frame, text="Stowage Code 엑셀 파일을 여기에 드롭하세요")
        self.stowage_drop_label.pack(pady=20)

        # 드래그 앤 드롭 바인딩
        self.stowage_drop_label.drop_target_register(DND_FILES)
        self.stowage_drop_label.dnd_bind('<<Drop>>', self.drop_stowage_file)

        # 현재 매핑 미리보기
        preview_frame = ttk.LabelFrame(main_frame, text="현재 매핑 미리보기")
        preview_frame.pack(fill="both", expand=True)
        
        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(preview_frame)
        scrollbar.pack(side="right", fill="y")
        
        # 미리보기 텍스트 위젯 생성
        self.preview_text = tk.Text(preview_frame, height=20, width=50, yscrollcommand=scrollbar.set)
        self.preview_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 스크롤바와 텍스트 위젯 연결
        scrollbar.config(command=self.preview_text.yview)

        # 저장 버튼
        save_button = ttk.Button(main_frame, text="설정 저장", command=self.save_stowage_settings)
        save_button.pack(pady=10)

        # 초기 미리보기 내용 설정
        preview_text = "=== 현재 매핑 ===\n"
        if self.stow_mapping:
            for service_name, mappings in self.stow_mapping.items():
                preview_text += f"\nService Name: {service_name}\n"
                preview_text += "------------------------\n"
                for mapping in mappings:
                    preview_text += f"Port: {mapping['port']}\n"
                    preview_text += f"Stow Code: {mapping['stow_code']}\n"
                    preview_text += "------------------------\n"
        else:
            preview_text += "매핑 정보가 없습니다."

        # 미리보기 텍스트 설정
        self.preview_text.delete(1.0, tk.END)
        self.preview_text.insert(tk.END, preview_text)

    def on_service_selected(self, event):
        """서비스 선택 시 처리"""
        self.update_stowage_preview()

    def on_entry_click(self, event, entry):
        """Entry 위젯 클릭시 기본 텍스트 제거"""
        if entry.get() in ["POD", "SzTp", "수량"]:
            entry.delete(0, tk.END)
            entry.config(foreground='black')

    def on_focus_out(self, event, entry, default_text):
        """Entry 위젯에서 포커스가 빠졌을 때 처리"""
        if entry.get().strip() == "":
            entry.insert(0, default_text)
            entry.config(foreground='gray')

    def on_key_press(self, event, entry):
        """키 입력 처리"""
        if entry.get() in ["POD", "SzTp", "수량"]:
            entry.delete(0, tk.END)

    def on_tab(self, event):
        """탭 키 처리"""
        current = event.widget
        next_widget = current.tk_focusNext()
        next_widget.focus()
        return "break"  # 기본 탭 동작 방지

    def drop_cll_file(self, event):
        """단일 CLL 파일 드롭 처리"""
        file_path = event.data.strip('{}').strip('"')
        if not os.path.exists(file_path):
            messagebox.showerror("오류", "파일이 존재하지 않습니다.")
            return

        try:
            # 엑셀 파일 읽기
            df_check = pd.read_excel(file_path, header=None)
            # 4행 12열의 값 가져오기 (0-based index이므로 3, 11)
            terminal_code = str(df_check.iloc[3, 11]).strip()

            if not terminal_code:
                messagebox.showerror("오류", "(4,12) 위치에서 터미널 코드를 찾을 수 없습니다.")
                return

            # 터미널 코드를 기반으로 POL, TOL 값 자동 설정
            port_info = self.terminal_to_port_mapping(terminal_code)
            
            if not port_info['pol'] or not port_info['tol']:
                messagebox.showerror("오류", f"터미널 코드 '{terminal_code}'에 대한 매핑을 찾을 수 없습니다.")
                return

            # POL, TOL 설정
            self.selected_pol.set(port_info['pol'])
            self.selected_tol.set(port_info['tol'])

            # 단일 탭의 POL 버튼 색상만 업데이트
            for port, btn in self.pol_buttons.items():
                if port == port_info['pol']:
                    btn.configure(bg='yellow')
                else:
                    btn.configure(bg='SystemButtonFace')

            # 단일 탭의 TOL 버튼 색상만 업데이트
            for terminal, btn in self.tol_buttons.items():
                if terminal == port_info['tol']:
                    btn.configure(bg='yellow')
                else:
                    btn.configure(bg='SystemButtonFace')

            # CLL 데이터 읽기
            df = pd.read_excel(file_path, header=4)
            
            # POD 목록 추출 및 매칭되는 서비스 찾기
            pod_list = df['POD'].unique().tolist()
            matching_services = self.find_matching_services(pod_list)
            
            if not matching_services:
                messagebox.showwarning("경고", "일치하는 서비스를 찾을 수 없습니다.")
                return
            
            # 서비스 선택 다이얼로그 표시
            selected_service = self.show_service_selection_dialog(matching_services)
            
            if not selected_service:
                return
                
            # 선택된 서비스 저장
            self.selected_service.set(selected_service)

            self.current_file = file_path
            self.input_label.config(text=f"입력 파일: {os.path.basename(file_path)}")
            
            # 단일 탭의 Summary 업데이트
            self.update_single_summary(df)
            
            # 멀티 탭의 Summary 초기화
            if hasattr(self, 'multi_summary_text'):
                self.multi_summary_text.delete(1.0, tk.END)
                self.multi_summary_text.insert(tk.END, "Multi CLL 탭에서 파일 병합 시 Summary가 표시됩니다.")
            
            # 파일 변환 실행
            self.convert_file()
            
        except Exception as e:
            error_msg = str(e)
            messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다:\n{error_msg}")

    def drop_obl_file(self, event):
        """OBL 파일 드롭 처리"""
        file_path = event.data
        file_path = file_path.strip('{}')
        self.current_file = file_path

        # 파일 정보 표시 업데이트
        self.input_label.config(text=f"입력 파일: {file_path}")
        self.obl_label.config(text=f"선택된 파일: {os.path.basename(file_path)}")

        # EMPTY 컨테이너 추가 실행
        self.add_empty_to_obl()

    def add_empty_to_obl(self):
        """기존 OBL에 EMPTY 컨테이너 추가"""
        # OBL 파일 읽기
        obl_df = pd.read_excel(self.current_file)

        # 기존 OBL의 컬럼 목록 가져오기
        existing_columns = obl_df.columns.tolist()

        # EMPTY 컨테이너 추가
        new_rows = []
        empty_container_num = 1  # 컨테이너 번호 시작값
        
        for pod_entry, sztp_entry, qty_entry in self.empty_entries:
            pod = pod_entry.get()
            sztp = sztp_entry.get()
            qty = qty_entry.get()

            if pod not in ["POD", ""] and sztp not in ["SzTp", ""] and qty not in ["수량", ""]:
                try:
                    qty = int(qty)
                    # SzTp를 정수로 변환
                    sztp = int(sztp)
                    
                    # SzTp에 따른 무게 설정
                    if str(sztp).startswith('2'):
                        weight = 2500
                    elif str(sztp).startswith('4'):
                        weight = 4500
                    else:
                        weight = 0

                    for i in range(qty):
                        # 기존 컬럼 구조를 따르는 빈 딕셔너리 생성
                        empty_row = {col: '' for col in existing_columns}

                        # 마지막 No 값 계산
                        last_no = len(obl_df) + len(new_rows) + 1

                        # EMPTY 컨테이너 번호 생성
                        ctr_nbr = f"MSCU{empty_container_num:07d}"
                        empty_container_num += 1

                        # 필요한 필드만 업데이트
                        empty_row.update({
                            'No': last_no,
                            'CtrNbr': ctr_nbr,  # 컨테이너 번호 설정
                            'ShOwn': 'N',
                            'Opr': 'MSC',
                            'POR': self.selected_pol.get(),
                            'POL': self.selected_pol.get(),
                            'TOL': self.selected_tol.get(),
                            'POD': pod,
                            'FPOD': pod,  # POD와 동일한 값으로 설정
                            'SzTp': sztp,
                            'Wgt': weight,  # SzTp에 따른 무게 설정
                            'ForE': 'E',
                            'Rfopr': 'N',
                            'Door': 'C',
                            'CustH': 'N',
                            'Fumi': 'N',
                            'VGM': 'Y',
                            'Stow': self.stow_mapping.get(pod, '')  # FPOD(POD)에 대한 Stow 코드
                        })
                        new_rows.append(empty_row)
                except ValueError:
                    continue  # 잘못된 입력은 조용히 건너뛰기

        # 새로운 EMPTY 컨테이너 추가
        if new_rows:
            new_df = pd.DataFrame(new_rows, columns=existing_columns)
            obl_df = pd.concat([obl_df, new_df], ignore_index=True)

            # 파일 저장
            input_dir = os.path.dirname(self.current_file)
            base_name = os.path.splitext(os.path.basename(self.current_file))[0]
            output_file = os.path.join(input_dir, f"{base_name}_EMPTY_ADDED.xlsx")
            obl_df.to_excel(output_file, index=False)

            self.output_file = output_file
            self.output_label.config(text=f"출력 파일: {output_file}")

            # Summary 업데이트
            self.update_summary(obl_df)

            messagebox.showinfo("성공", "EMPTY 컨테이너가 추가되었습니다.")

    def update_summary(self, df):
        """컨테이너 요약 정보 업데이트"""
        summary = "=== FULL 컨테이너 ===\n"
        full_containers = df[df['F/E'] == 'F']
        full_summary = full_containers['T&S'].value_counts()
        for sztp, count in full_summary.items():
            summary += f"{sztp}: {count}개\n"
        summary += f"FULL 컨테이너 총계: {len(full_containers)}개\n"

        summary += "\n=== EMPTY 컨테이너 ===\n"
        empty_containers = df[df['F/E'] == 'E']
        empty_summary = empty_containers['T&S'].value_counts()
        for sztp, count in empty_summary.items():
            summary += f"{sztp}: {count}개\n"

        # EMPTY 입력란에서 추가될 컨테이너 계산
        additional_empty = 0
        for pod_entry, sztp_entry, qty_entry in self.empty_entries:
            qty = qty_entry.get()
            if qty not in ["수량", ""]:
                try:
                    additional_empty += int(qty)
                except ValueError:
                    pass

        total_empty = len(empty_containers) + additional_empty
        summary += f"EMPTY 컨테이너 총계: {total_empty}개\n"

        # 전체 총계
        summary += f"\n=== 전체 컨테이너 ===\n"
        summary += f"총계: {len(full_containers) + total_empty}개"

        self.summary_text.delete(1.0, tk.END)
        self.summary_text.insert(tk.END, summary)

    def select_pol(self, port):
        """POL 버튼 선택 처리"""
        self.selected_pol.set(port)
        # 모든 버튼 원래 색으로
        for btn in self.pol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        # 선택된 버튼만 노란색으로
        self.pol_buttons[port].configure(bg='yellow')

    def select_tol(self, terminal):
        """TOL 버튼 선택 처리"""
        self.selected_tol.set(terminal)
        # 모든 버튼 원래 색으로
        for btn in self.tol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        # 선택된 버튼만 노란색으로
        self.tol_buttons[terminal].configure(bg='yellow')

    def convert_file(self):
        """단일 CLL 파일 변환"""
        try:
            # 선택된 서비스 확인
            selected_service = self.selected_service.get()
            if not selected_service:
                messagebox.showwarning("경고", "Service Name을 선택해주세요!")
                return

            # CLL 파일 읽기
            cll_df = pd.read_excel(self.current_file, header=4)

            # 선택된 서비스의 매핑 가져오기
            service_mappings = self.stow_mapping.get(selected_service, [])

            # OBL 데이터프레임 생성
            obl_data = []

            # CLL 데이터 변환
            for idx, row in cll_df.iterrows():
                # OPT가 비어있으면 선택된 POL 값 사용
                por_value = row['OPT'] if pd.notna(row['OPT']) and row['OPT'] != '' else self.selected_pol.get()

                # POD와 FPOD 처리
                pod = str(row['POD']) if pd.notna(row['POD']) else ''
                fpod = str(row['FDP']) if pd.notna(row['FDP']) else ''  # FPOD는 CLL의 FDP 값 사용
                
                # 초기값 설정
                mapped_port = pod  # POD 초기값
                mapped_stow = ''   # Stow 초기값
                
                # POD와 FPOD가 다른 경우에만 stow_code 매핑 적용
                if pod != fpod:
                    # POD가 stow_code와 일치하는지 확인
                    for mapping in service_mappings:
                        if pod.upper() == mapping['stow_code'].upper():
                            mapped_port = mapping['port']      # POD를 port 값으로 설정
                            mapped_stow = mapping['stow_code'] # Stow를 stow_code 값으로 설정
                            break

                obl_row = {
                    'No': idx + 1,
                    'CtrNbr': row['CNTR NO'],
                    'ShOwn': 'N',
                    'Opr': 'MSC',
                    'POR': por_value,
                    'POL': self.selected_pol.get(),
                    'TOL': self.selected_tol.get(),
                    'POD': mapped_port,
                    'TOD': '',
                    'Stow': mapped_stow,
                    'FPOD': fpod,  # FPOD는 원래 값 유지
                    'SzTp': int(row['T&S']) if pd.notna(row['T&S']) else '',
                    'Wgt': int(row['WGT']) if pd.notna(row['WGT']) else '',
                    'ForE': row['F/E'],
                    'Lbl': '',
                    'Rfopr': 'N',
                    'Rftemp': row['R/F'].replace(' CEL', '') if pd.notna(row['R/F']) else '',
                    'OvDH': row['OH'],
                    'OvDF': row['OL'] / 2 if pd.notna(row['OL']) and row['OL'] != 0 else '',
                    'OvDA': row['OL'] / 2 if pd.notna(row['OL']) and row['OL'] != 0 else '',
                    'OvDP': row['OW'] / 2 if pd.notna(row['OW']) and row['OW'] != 0 else '',
                    'OvDS': row['OW'] / 2 if pd.notna(row['OW']) and row['OW'] != 0 else '',
                    'OvSH': '',
                    'OvSF': '',
                    'OvSA': '',
                    'OvSP': '',
                    'OvSS': '',
                    'BL': '',
                    'HI': '',
                    'AC': '',
                    'Flip': '',
                    'Door': 'C',
                    'CustH': 'N',
                    'LenBB': '',
                    'BrthBB': '',
                    'HgtBB': '',
                    'WgtBB': '',
                    'Fumi': 'N',
                    'FuDt': '',
                    'VenDt': '',
                    'Venti': '',
                    'Damag': '',
                    'PPK': '',
                    'Food': '',
                    'Resi': '',
                    'Book': '',
                    'Cold': '',
                    'Catm': '',
                    'VGM': 'Y',
                    'VGM Weighting Method': '',
                    'HVC': '',
                    'BN1': '',
                    'BN2': '',
                    'BN3': '',
                    'BN4': '',
                    'Harmonised system codes': '',
                    'Description': '',
                    'Flexitank': '',
                    'UNNO': row['UNDG'],
                    'Class': row['IMDG'],
                    'PSN': '',
                    'N.Weight': '',
                    'S.Risk1': '',
                    'S.Risk2': '',
                    'S.Risk3': '',
                    'P.Group': '',
                    'LQ': '',
                    'EQ': '',
                    'FP': '',
                    'IMDG Remark': '',
                    'Sub Index': '',
                    'Inf type': '',
                    'Address': '',
                    'Street': '',
                    'City': '',
                    'Postal Code': '',
                    'Country Code': '',
                    'Country': '',
                    'Sub Index_1': '',  # 첫 번째 Sub Index 열
                    'Remark': ''
                }
                obl_data.append(obl_row)

            # EMPTY 컨테이너 추가 로직
            last_no = len(obl_data)
            empty_container_num = 1
            for pod_entry, sztp_entry, qty_entry in self.empty_entries:
                pod = pod_entry.get()
                sztp = sztp_entry.get()
                qty = qty_entry.get()

                if pod not in ["POD", ""] and sztp not in ["SzTp", ""] and qty not in ["수량", ""]:
                    try:
                        qty = int(qty)
                        for i in range(qty):
                            empty_row = dict.fromkeys(obl_data[0].keys(), '')
                            
                            # POD에 대한 매핑 확인
                            mapped_port = pod
                            mapped_stow = ''
                            
                            # POD가 stow_code와 일치하는지 확인
                            for mapping in service_mappings:
                                if pod.upper() == mapping['stow_code'].upper():
                                    # stow_code가 일치하면 해당 port를 POD로 사용
                                    mapped_port = mapping['port']
                                    mapped_stow = mapping['stow_code']
                                    break
                                elif pod.upper() == mapping['port'].upper():
                                    # port가 일치하면 해당 stow_code 사용
                                    mapped_port = mapping['port']
                                    mapped_stow = mapping['stow_code']
                                    break
                            
                            empty_row.update({
                                'No': last_no + 1,
                                'CtrNbr': f"MSCU{empty_container_num:07d}",
                                'ShOwn': 'N',
                                'Opr': 'MSC',
                                'POR': self.selected_pol.get(),
                                'POL': self.selected_pol.get(),
                                'TOL': self.selected_tol.get(),
                                'POD': mapped_port,
                                'FPOD': mapped_port,
                                'SzTp': int(sztp),
                                'Wgt': int(2500 if str(sztp).startswith('2') else 4700 if str(sztp).startswith('4') else 0),
                                'ForE': 'E',
                                'Rfopr': 'N',
                                'Door': 'C',
                                'CustH': 'N',
                                'Fumi': 'N',
                                'VGM': 'Y',
                                'Stow': mapped_stow
                            })
                            obl_data.append(empty_row)
                            last_no += 1
                            empty_container_num += 1
                    except ValueError:
                        messagebox.showwarning("경고", f"잘못된 수량 형식: {qty}")

            # OBL 데이터프레임 생성 s
            obl_df = pd.DataFrame(obl_data)

            # 파일 저장
            input_dir = os.path.dirname(self.current_file)
            base_name = os.path.splitext(os.path.basename(self.current_file))[0]
            # 엑셀 파일에서 vessel name과 항차, 터미널 정보 읽기
            # 현재 파일에서 vessel name, voyage no, terminal 정보 읽기
            current_df = pd.read_excel(self.current_file)
            vessel_name = str(current_df.iloc[1, 2]).strip()  # C3 셀의 vessel name
            voyage_no = str(current_df.iloc[1, 11]).strip()   # L3 셀의 항차 정보
            terminal = str(current_df.iloc[2, 11]).strip()    # L4 셀의 터미널 정보
            
            output_file = os.path.join(input_dir, f"{base_name}_{vessel_name}_{voyage_no}_{terminal}_OBL.xlsx")
            obl_df = obl_df.rename(columns={
                'Sub Index_1': 'Sub Index'
            })
            obl_df.to_excel(output_file, index=False)

            self.output_file = output_file
            self.output_label.config(text=f"출력 파일: {output_file}")

            # 단일 탭의 Summary만 업데이트
            self.update_single_summary(cll_df)
            
            messagebox.showinfo("성공", "변환이 완료되었습니다.")

        except Exception as e:
            messagebox.showerror("Error", f"변환 중 오류 발생: {str(e)}")

    def update_single_summary(self, df):
        """단일 CLL 파일의 Container Summary 업데이트"""
        try:
            self.single_summary_text.delete(1.0, tk.END)
            
            summary_text = "=== 단일 CLL 변환 Summary ===\n"
            summary_text += "================================\n\n"
            
            # 데이터프레임 유효성 검사
            if df is None or df.empty:
                raise ValueError("유효한 데이터가 없습니다.")
            
            # 컬럼 존재 여부 확인
            required_columns = ['T&S', 'F/E', 'POD']
            for col in required_columns:
                if col not in df.columns:
                    raise ValueError(f"필요한 컬럼이 없습니다: {col}")
            
            # 데이터 처리
            total_containers = len(df)
            
            # 각 컬럼별 카운트 계산 (NaN 값 제외)
            size_type_counts = df['T&S'].dropna().value_counts()
            full_empty_counts = df['F/E'].dropna().value_counts()
            pod_counts = df['POD'].dropna().value_counts()
            
            # Summary 텍스트 구성
            summary_text += f"Total Containers: {total_containers}\n"
            summary_text += "--------------------------------\n\n"
            
            summary_text += "=== Size Type Summary ===\n"
            for sz_tp, count in size_type_counts.items():
                if pd.notna(sz_tp):  # NaN 값 체크
                    summary_text += f"{sz_tp}: {count}\n"
            summary_text += "--------------------------------\n\n"
            
            summary_text += "=== Full/Empty Summary ===\n"
            for fe, count in full_empty_counts.items():
                if pd.notna(fe):  # NaN 값 체크
                    summary_text += f"{fe}: {count}\n"
            summary_text += "--------------------------------\n\n"
            
            summary_text += "=== POD Summary ===\n"
            for pod, count in pod_counts.items():
                if pd.notna(pod):  # NaN 값 체크
                    summary_text += f"{pod}: {count}\n"
            summary_text += "--------------------------------"
            
            self.single_summary_text.insert(tk.END, summary_text)
            
        except Exception as e:
            print(f"Summary 생성 중 오류 발생: {str(e)}")  # 디버깅용
            self.single_summary_text.delete(1.0, tk.END)
            self.single_summary_text.insert(tk.END, "단일 CLL 탭에서 파일 변환 시 Summary가 표시됩니다.")

    def drop_master_cll(self, event):
        """Master CLL 파일 드롭 처리"""
        file_path = event.data.strip('{}').strip('"')
        if not os.path.exists(file_path):
            messagebox.showerror("오류", "파일이 존재하지 않습니다.")
            return

        try:
            # 엑셀 파일 읽기
            df_check = pd.read_excel(file_path, header=None)
            # 4행 12열의 값 가져오기 (0-based index이므로 3, 11)
            terminal_code = str(df_check.iloc[3, 11]).strip()

            if not terminal_code:
                messagebox.showerror("오류", "(4,12) 위치에서 터미널 코드를 찾을 수 없습니다.")
                return

            # 터미널 코드를 기반으로 POL, TOL 값 자동 설정
            port_info = self.terminal_to_port_mapping(terminal_code)
            
            if not port_info['pol'] or not port_info['tol']:
                messagebox.showerror("오류", f"터미널 코드 '{terminal_code}'에 대한 매핑을 찾을 수 없습니다.")
                return

            # POL, TOL 설정
            self.selected_pol.set(port_info['pol'])
            self.selected_tol.set(port_info['tol'])

            # Multi 탭의 POL 버튼 색상만 업데이트
            for port, btn in self.multi_pol_buttons.items():
                if port == port_info['pol']:
                    btn.configure(bg='yellow')
                else:
                    btn.configure(bg='SystemButtonFace')

            # Multi 탭의 TOL 버튼 색상만 업데이트
            for btn_text, value in self.tol_values.items():
                if value == port_info['tol']:
                    self.multi_tol_buttons[btn_text].configure(bg='yellow')
                else:
                    self.multi_tol_buttons[btn_text].configure(bg='SystemButtonFace')

            self.master_file = file_path
            self.master_path_label.config(text=f"파일 경로: {file_path}")
            self.master_label.config(text="Master 파일이 선택되었습니다")
            
            # Slave 프레임 활성화
            self.slave_frame.pack(pady=10, padx=10, fill="x")

        except Exception as e:
            messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")

    def drop_slave_cll(self, event):
        """Slave CLL 파일 드롭 처리"""
        if not hasattr(self, 'master_file'):
            messagebox.showwarning("경고", "Master 파일을 먼저 선택해주세요!")
            return

        file_path = event.data.strip('{}').strip('"')
        if not os.path.exists(file_path):
            messagebox.showerror("오류", "파일이 존재하지 않습니다.")
            return

        self.slave_file = file_path
        self.slave_path_label.config(text=f"파일 경로: {file_path}")
        self.slave_label.config(text="Slave 파일이 선택되었습니다")
        
        # Slave 파일이 선택되면 바로 병합 처리 시작
        self.combine_cll_files()

    def select_multi_pol(self, port):
        """Multi 탭 POL 버튼 선택 처리"""
        self.selected_pol.set(port)
        # 모든 버튼 원래 색으로
        for btn in self.multi_pol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        # 선택된 버튼만 노란색으로
        self.multi_pol_buttons[port].configure(bg='yellow')

    def select_multi_tol(self, terminal):
        """Multi 탭 TOL 버튼 선택 처리"""
        self.selected_tol.set(terminal)
        # 모든 버튼 원래 색으로
        for btn in self.multi_tol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        # 선택된 버튼만 노란색으로
        for btn_text, value in self.tol_values.items():
            if value == terminal:
                self.multi_tol_buttons[btn_text].configure(bg='yellow')

    def combine_cll_files(self):
        """Master와 Slave CLL 파일 병합"""
        try:
            # Master와 Slave 파일의 POD 목록 추출
            master_df = pd.read_excel(self.master_file, header=4)
            slave_df = pd.read_excel(self.slave_file, header=4)
            
            # 두 파일의 POD 목록 합치기
            pod_list = list(set(master_df['POD'].unique().tolist() + slave_df['POD'].unique().tolist()))
            
            # 매칭되는 서비스 찾기
            matching_services = self.find_matching_services(pod_list)
            
            if not matching_services:
                messagebox.showwarning("경고", "일치하는 서비스를 찾을 수 없습니다.")
                return
            
            # 서비스 선택 다이얼로그 표시
            selected_service = self.show_service_selection_dialog(matching_services)
            
            if not selected_service:
                return
                
            # 선택된 서비스 저장
            self.selected_service.set(selected_service)

            def process_cll_file(file_path, start_row):
                cll_df = pd.read_excel(file_path, header=4)
                processed_data = []
                row_count = start_row

                # 선택된 서비스의 매핑 가져오기
                service_mappings = self.stow_mapping.get(selected_service, [])

                for idx, row in cll_df.iterrows():
                    if pd.notna(row['CNTR NO']):
                        # POD 값 가져오기
                        pod = str(row['POD']) if pd.notna(row['POD']) else ''
                        fpod = str(row['FDP']) if pd.notna(row['FDP']) else ''
                        
                        # 초기값 설정
                        mapped_port = pod
                        mapped_stow = ''
                        
                        # POD가 stow_code와 일치하는지 확인
                        for mapping in service_mappings:
                            if pod.upper() == mapping['stow_code'].upper():
                                mapped_port = mapping['port']
                                mapped_stow = mapping['stow_code']
                                break
                            elif pod.upper() == mapping['port'].upper():
                                mapped_port = mapping['port']
                                mapped_stow = mapping['stow_code']
                                break

                        obl_row = {
                            'No': row_count,
                            'CtrNbr': row['CNTR NO'],
                            'ShOwn': 'N',
                            'Opr': 'MSC',
                            'POR': row['OPT'] if pd.notna(row['OPT']) else self.selected_pol.get(),
                            'POL': self.selected_pol.get(),
                            'TOL': self.selected_tol.get(),
                            'POD': mapped_port,
                            'TOD': '',
                            'Stow': mapped_stow,
                            'FPOD': fpod,
                            'SzTp': int(row['T&S']) if pd.notna(row['T&S']) else '',
                            'Wgt': int(row['WGT']) if pd.notna(row['WGT']) else '',
                            'ForE': row['F/E'],
                            'Lbl': '',
                            'Rfopr': 'N',
                            'Rftemp': row['R/F'].replace(' CEL', '') if pd.notna(row['R/F']) else '',
                            'OvDH': row['OH'],
                            'OvDF': row['OL'] / 2 if pd.notna(row['OL']) and row['OL'] != 0 else '',
                            'OvDA': row['OL'] / 2 if pd.notna(row['OL']) and row['OL'] != 0 else '',
                            'OvDP': row['OW'] / 2 if pd.notna(row['OW']) and row['OW'] != 0 else '',
                            'OvDS': row['OW'] / 2 if pd.notna(row['OW']) and row['OW'] != 0 else '',
                            'OvSH': '',
                            'OvSF': '',
                            'OvSA': '',
                            'OvSP': '',
                            'OvSS': '',
                            'BL': '',
                            'HI': '',
                            'AC': '',
                            'Flip': '',
                            'Door': 'C',
                            'CustH': 'N',
                            'LenBB': '',
                            'BrthBB': '',
                            'HgtBB': '',
                            'WgtBB': '',
                            'Fumi': 'N',
                            'FuDt': '',
                            'VenDt': '',
                            'Venti': '',
                            'Damag': '',
                            'PPK': '',
                            'Food': '',
                            'Resi': '',
                            'Book': '',
                            'Cold': '',
                            'Catm': '',
                            'VGM': 'Y',
                            'VGM Weighting Method': '',
                            'HVC': '',
                            'BN1': '',
                            'BN2': '',
                            'BN3': '',
                            'BN4': '',
                            'Harmonised system codes': '',
                            'Description': '',
                            'Flexitank': '',
                            'UNNO': row['UNDG'],
                            'Class': row['IMDG'],
                            'PSN': '',
                            'N.Weight': '',
                            'S.Risk1': '',
                            'S.Risk2': '',
                            'S.Risk3': '',
                            'P.Group': '',
                            'LQ': '',
                            'EQ': '',
                            'FP': '',
                            'IMDG Remark': '',
                            'Sub Index': '',
                            'Inf type': '',
                            'Address': '',
                            'Street': '',
                            'City': '',
                            'Postal Code': '',
                            'Country Code': '',
                            'Country': '',
                            'Sub Index_1': '',  # 첫 번째 Sub Index 열
                            'Remark': ''
                        }
                        processed_data.append(obl_row)
                        row_count += 1
                return processed_data

            # Master와 Slave 파일 처리
            master_data = process_cll_file(self.master_file, 1)
            slave_data = process_cll_file(self.slave_file, len(master_data) + 1)
            
            all_data = master_data + slave_data
            
            # DataFrame 생성 및 저장
            combined_df = pd.DataFrame(all_data)
            save_dir = os.path.dirname(self.master_file)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(save_dir, f"Combined_OBL_{timestamp}.xlsx")
            
            # 파일 저장
            combined_df.to_excel(output_file, index=False)
            
            # 결과 표시
            self.result_label.config(text=f"출력 파일: {output_file}")
            
            # 멀티 탭의 Summary만 업데이트
            self.update_multi_summary(combined_df)
            
            # 단일 탭의 Summary는 초기화
            if hasattr(self, 'single_summary_text'):
                self.single_summary_text.delete(1.0, tk.END)
                self.single_summary_text.insert(tk.END, "단일 CLL 탭에서 파일 변환 시 Summary가 표시됩니다.")
            
            messagebox.showinfo("성공", f"CLL 파일들이 성공적으로 병합되었습니다.\n총 {len(all_data)}개의 컨테이너가 처리되었습니다.")
            
        except Exception as e:
            print(f"Error in combine_cll_files: {str(e)}")  # 디버깅용
            messagebox.showerror("오류", str(e))

    def update_multi_summary(self, df):
        """병합된 CLL 파일들의 Container Summary 업데이트"""
        try:
            self.multi_summary_text.delete(1.0, tk.END)
            
            summary_text = "=== CLL 병합 Summary ===\n"
            summary_text += "================================\n\n"
            
            # 데이터프레임 유효성 검사
            if df is None or df.empty:
                raise ValueError("유효한 데이터가 없습니다.")
            
            # 컬럼 존재 여부 확인
            required_columns = ['SzTp', 'ForE', 'POD']
            for col in required_columns:
                if col not in df.columns:
                    raise ValueError(f"필요한 컬럼이 없습니다: {col}")
            
            # 데이터 처리
            total_containers = len(df)
            
            # 각 컬럼별 카운트 계산 (NaN 값 제외)
            size_type_counts = df['SzTp'].dropna().value_counts()
            full_empty_counts = df['ForE'].dropna().value_counts()
            pod_counts = df['POD'].dropna().value_counts()
            
            # Summary 텍스트 구성
            summary_text += f"Total Containers: {total_containers}\n"
            summary_text += "--------------------------------\n\n"
            
            summary_text += "=== Size Type Summary ===\n"
            for sz_tp, count in size_type_counts.items():
                if pd.notna(sz_tp):
                    summary_text += f"{sz_tp}: {count}\n"
            summary_text += "--------------------------------\n\n"
            
            summary_text += "=== Full/Empty Summary ===\n"
            for fe, count in full_empty_counts.items():
                if pd.notna(fe):
                    summary_text += f"{fe}: {count}\n"
            summary_text += "--------------------------------\n\n"
            
            summary_text += "=== POD Summary ===\n"
            for pod, count in pod_counts.items():
                if pd.notna(pod):
                    summary_text += f"{pod}: {count}\n"
            summary_text += "--------------------------------"
            
            self.multi_summary_text.insert(tk.END, summary_text)
            
        except Exception as e:
            self.multi_summary_text.delete(1.0, tk.END)
            self.multi_summary_text.insert(tk.END, f"Summary 생성 중 오류 발생: {str(e)}")

    def reset_all(self):
        """프로그램 상태 초기화"""
        # POL/TOL 버튼 초기화 (단일 탭)
        for btn in self.pol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        for btn in self.tol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        
        # POL/TOL 버튼 초기화 (멀티 탭)
        for btn in self.multi_pol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        for btn in self.multi_tol_buttons.values():
            btn.configure(bg='SystemButtonFace')
        
        # 선택값 초기화
        self.selected_pol.set('')
        self.selected_tol.set('')
        
        # 파일 경로 레이블 초기화
        self.input_label.config(text="입력 파일: 없음")
        self.output_label.config(text="출력 파일: 없음")
        self.master_path_label.config(text="파일 경로: 없음")
        self.slave_path_label.config(text="파일 경로: 없음")
        self.result_label.config(text="출력 파일: 없음")
        
        # Summary 텍스트 초기화
        self.single_summary_text.delete(1.0, tk.END)
        self.single_summary_text.insert(tk.END, "단일 CLL 탭에서 파일 변환 시 Summary가 표시됩니다.")
        self.multi_summary_text.delete(1.0, tk.END)
        self.multi_summary_text.insert(tk.END, "Multi CLL 탭에서 파일 변환 시 Summary가 표시됩니다.")
        
        # 파일 관련 변수 초기화
        self.current_file = None
        self.output_file = None
        if hasattr(self, 'master_file'):
            delattr(self, 'master_file')
        if hasattr(self, 'slave_file'):
            delattr(self, 'slave_file')

        # Entry 위젯 초기화
        for pod_entry, sztp_entry, qty_entry in self.empty_entries:
            # Entry 위젯 상태 초기화
            pod_entry.delete(0, tk.END)
            sztp_entry.delete(0, tk.END)
            qty_entry.delete(0, tk.END)
            
            # 플레이스홀더 텍스트 설정
            pod_entry.insert(0, "POD")
            sztp_entry.insert(0, "SzTp")
            qty_entry.insert(0, "수량")
            
            # Entry 위젯 상태 설정
            pod_entry.config(state='normal')
            sztp_entry.config(state='normal')
            qty_entry.config(state='normal')

    def drop_itps_file(self, event):
        """ITPS 파일 드롭 처리"""
        file_path = event.data.strip('{}').strip('"')
        if not os.path.exists(file_path):
            messagebox.showerror("오류", "파일이 존재하지 않습니다.")
            return

        self.itps_file = file_path
        self.itps_input_label.config(text=f"ITPS 파일: {os.path.basename(file_path)}")
        self.itps_drop_label.config(text="ITPS 파일이 선택되었습니다")
        
        # 두 파일이 모두 선택되었다면 자동으로 처리 시작
        if self.itps_file and self.obl_file:
            self.process_itps_file()

    def drop_obl_for_itps(self, event):
        """ITPS 처리를 위한 OBL 파일 드롭 처리"""
        file_path = event.data.strip('{}').strip('"')
        if not os.path.exists(file_path):
            messagebox.showerror("오류", "파일이 존재하지 않습니다.")
            return

        self.obl_file = file_path
        self.itps_obl_label.config(text=f"OBL 파일: {os.path.basename(file_path)}")
        self.obl_drop_label.config(text="OBL 파일이 선택되었습니다")
        
        # 두 파일이 모두 선택되었다면 자동으로 처리 시작
        if self.itps_file and self.obl_file:
            self.process_itps_file()

    def process_itps_file(self):
        """ITPS 파일 처리 및 OBL에 추가"""
        try:
            print("Starting ITPS file processing...")  # 디버깅용

            # OBL 파일 읽기
            print(f"Reading OBL file: {self.obl_file}")  # 디버깅용
            obl_df = pd.read_excel(self.obl_file)
            print(f"OBL data rows: {len(obl_df)}")  # 디버깅용
            
            # ITPS 파일 읽기
            print(f"Reading ITPS file: {self.itps_file}")  # 디버깅용
            itps_df = pd.read_excel(self.itps_file)
            print(f"ITPS data rows: {len(itps_df)}")  # 디버깅용
            
            # 기존 OBL의 마지막 No 값 가져오기
            last_no = len(obl_df)
            
            # OBL의 POL과 TOL 값 가져오기
            obl_pol = obl_df['POL'].iloc[0] if not obl_df.empty else ''
            obl_tol = obl_df['TOL'].iloc[0] if not obl_df.empty else ''

            # 새로운 행들을 저장할 리스트
            new_rows = []

            # ITPS 데이터를 OBL 형식으로 변환
            for idx, row in itps_df.iterrows():
                try:
                    # Equipment Number가 있는 경우만 처리
                    if pd.isna(row['Equipment Number']):
                        continue

                    print(f"Processing ITPS row {idx}: {row['Equipment Number']}")  # 디버깅용
                    
                    # PORT CODE 변환
                    por = self.convert_to_port_code(row['Origin Load Port']) if pd.notna(row['Origin Load Port']) else ''
                    pol = self.convert_to_port_code(obl_pol)
                    pod = self.convert_to_port_code(row['Discharge Port']) if pd.notna(row['Discharge Port']) else ''
                    
                    # TpSZ 매핑
                    tpsz = str(row['Type/Size']) if pd.notna(row['Type/Size']) else ''
                    mapped_tpsz = self.tpsz_mapping.get(tpsz, tpsz)
                    
                    # Rftemp 처리
                    rftemp = ''
                    if pd.notna(row['Reefer Temp.']):
                        temp_str = str(row['Reefer Temp.']).split('/')[0].strip()
                        rftemp = temp_str

                    # Weight 처리
                    weight = ''
                    if pd.notna(row['Weight']):
                        try:
                            weight = int(float(row['Weight']))
                        except:
                            weight = ''

                    # 새로운 행 데이터 생성
                    new_row = pd.Series(index=obl_df.columns)  # OBL의 모든 컬럼으로 초기화
                    new_row.fillna('', inplace=True)  # 모든 값을 빈 문자열로 초기화

                    # 데이터 매핑
                    new_row['No'] = last_no + len(new_rows) + 1
                    new_row['CtrNbr'] = str(row['Equipment Number'])
                    new_row['ShOwn'] = 'N'
                    new_row['Opr'] = 'MSC'
                    new_row['POR'] = por
                    new_row['POL'] = pol
                    new_row['TOL'] = obl_tol
                    new_row['POD'] = pod
                    new_row['FPOD'] = pod
                    new_row['SzTp'] = mapped_tpsz
                    new_row['Wgt'] = weight
                    new_row['ForE'] = str(row['Full/Empty']) if pd.notna(row['Full/Empty']) else ''
                    new_row['Rfopr'] = 'N'
                    new_row['Rftemp'] = rftemp
                    new_row['Door'] = 'C'
                    new_row['CustH'] = 'N'
                    new_row['Fumi'] = 'N'
                    new_row['VGM'] = 'Y'
                    
                    # IMO Class 처리
                    if pd.notna(row['IMO Class']):
                        imo_class = str(row['IMO Class'])
                        if imo_class.replace('.', '').isdigit():
                            new_row['Class'] = str(int(float(imo_class)))
                        else:
                            new_row['Class'] = imo_class

                    # UN Number 처리
                    if pd.notna(row['UN Number']):
                        new_row['UNNO'] = str(row['UN Number'])[:6]

                    new_rows.append(new_row)
                    print(f"Added new row for container: {new_row['CtrNbr']}")  # 디버깅용
                    
                except Exception as e:
                    print(f"Error processing row {idx}: {str(e)}")  # 디버깅용
                    continue

            print(f"Total new rows created: {len(new_rows)}")  # 디버깅용

            # 새로운 데이터를 DataFrame으로 변환
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                print(f"New DataFrame created with {len(new_df)} rows")  # 디버깅용

                # 기존 OBL 데이터와 새로운 데이터 결합
                combined_df = pd.concat([obl_df, new_df], ignore_index=True)
                print(f"Combined DataFrame has {len(combined_df)} rows")  # 디버깅용

                # 파일 저장
                save_dir = os.path.dirname(self.obl_file)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = os.path.join(save_dir, f"OBL_with_ITPS_{timestamp}.xlsx")
                
                # 파일 저장 전 확인
                print(f"Saving to: {output_file}")  # 디버깅용
                combined_df.to_excel(output_file, index=False)
                
                # 파일 저장 확인
                if os.path.exists(output_file):
                    print(f"File saved successfully: {output_file}")  # 디버깅용
                    self.itps_output_label.config(text=f"출력 파일: {os.path.basename(output_file)}")
                    self.update_itps_summary(combined_df)
                    messagebox.showinfo("성공", "ITPS 데이터가 성공적으로 추가되었습니다.")
                else:
                    raise Exception("파일이 생성되지 않았습니다.")

            else:
                raise Exception("처리할 ITPS 데이터가 없습니다.")
            
        except Exception as e:
            print(f"Error in process_itps_file: {str(e)}")  # 디버깅용
            messagebox.showerror("오류", f"ITPS 처리 중 오류 발생: {str(e)}")

    def update_itps_summary(self, df):
        """ITPS 처리 결과 Summary 업데이트"""
        try:
            self.itps_summary_text.delete(1.0, tk.END)
            
            summary_text = "=== ITPS 추가 결과 Summary ===\n"
            summary_text += "================================\n\n"
            
            # 전체 컨테이너 수
            total_containers = len(df)
            summary_text += f"전체 컨테이너 수: {total_containers}\n"
            summary_text += "--------------------------------\n\n"
            
            # F/E 별 통계
            fe_counts = df['ForE'].value_counts()
            summary_text += "=== Full/Empty 현황 ===\n"
            for fe, count in fe_counts.items():
                summary_text += f"{fe}: {count}개\n"
            summary_text += "--------------------------------\n\n"
            
            # Size Type 별 통계
            sztp_counts = df['SzTp'].value_counts()
            summary_text += "=== Size Type 현황 ===\n"
            for sztp, count in sztp_counts.items():
                if pd.notna(sztp):
                    summary_text += f"{sztp}: {count}개\n"
            summary_text += "--------------------------------\n\n"
            
            # POD 별 통계
            pod_counts = df['POD'].value_counts()
            summary_text += "=== POD 현황 ===\n"
            for pod, count in pod_counts.items():
                if pd.notna(pod):
                    summary_text += f"{pod}: {count}개\n"
            summary_text += "--------------------------------"
            
            self.itps_summary_text.insert(tk.END, summary_text)
            
        except Exception as e:
            self.itps_summary_text.delete(1.0, tk.END)
            self.itps_summary_text.insert(tk.END, f"Summary 생성 중 오류 발생: {str(e)}")

    def convert_to_port_code(self, port_name):
        """항구 이름을 5자리 PORT CODE로 변환"""
        if not port_name or pd.isna(port_name):
            return ''
            
        port_name = str(port_name).strip().upper()
        
        # 이미 5자리 코드인 경우 그대로 반환
        if len(port_name) == 5 and port_name.isalnum():
            return port_name
            
        # port_codes의 value(port name)와 매칭 시도
        for code, full_name in self.port_codes.items():
            if full_name == port_name:  # 정확한 매칭
                return code
            elif full_name in port_name or port_name in full_name:  # 부분 매칭
                return code
                
        # 매칭되는 코드가 없으면 원래 값 반환
        return port_name

    def drop_stowage_file(self, event):
        """Stowage Code 엑셀 파일 드롭 처리"""
        try:
            file_path = event.data.strip('{}').strip('"')
            if not os.path.exists(file_path):
                messagebox.showerror("오류", "파일이 존재하지 않습니다.")
                return

            # 엑셀 파일 읽기 (헤더는 2번째 행, 데이터는 3번째 행부터)
            df = pd.read_excel(file_path, header=1)
            
            # 매핑 딕셔너리 생성
            service_mappings = {}
            for _, row in df.iterrows():
                service_name = str(row['Service Name']).strip()
                stow_code = str(row['Stow Code OBL7']).strip()
                
                # Port 열에서 [ ] 안의 값 추출
                port_str = str(row['Port']).strip()
                port = ''
                if '[' in port_str and ']' in port_str:
                    start = port_str.find('[') + 1
                    end = port_str.find(']')
                    port = port_str[start:end].strip()
                
                if port and stow_code:  # port와 stow_code가 있는 경우만 매핑에 추가
                    if service_name not in service_mappings:
                        service_mappings[service_name] = []
                    service_mappings[service_name].append({
                        'port': port,
                        'stow_code': stow_code
                    })
            
            # 설정 저장
            self.stow_mapping = service_mappings
            
            # 엑셀 파일 경로 저장
            excel_dir = os.path.dirname(file_path)
            excel_name = os.path.splitext(os.path.basename(file_path))[0]
            self.stowage_config_file = os.path.join(excel_dir, f"{excel_name}_mapping.json")
            
            self.save_stowage_settings()
            
            # 미리보기 업데이트
            self.update_stowage_preview()
            
            messagebox.showinfo("성공", "Stowage Code 매핑이 성공적으로 업데이트되었습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")

    def save_stowage_settings(self):
        """Stowage Code 설정 저장"""
        try:
            # JSON 파일로 저장
            with open(self.stowage_config_file, 'w', encoding='utf-8') as f:
                json.dump(self.stow_mapping, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("성공", f"설정이 성공적으로 저장되었습니다.\n저장 위치: {self.stowage_config_file}")
            
        except Exception as e:
            messagebox.showerror("오류", f"설정 저장 중 오류가 발생했습니다: {str(e)}")

    def update_stowage_preview(self):
        """Stowage Code 매핑 미리보기 업데이트"""
        try:
            if hasattr(self, 'preview_text'):
                self.preview_text.delete(1.0, tk.END)
                
                preview_text = "=== 현재 매핑 ===\n"
                
                if not self.stow_mapping:
                    preview_text += "매핑 정보가 없습니다."
                else:
                    for service_name, mappings in self.stow_mapping.items():
                        preview_text += f"\nService Name: {service_name}\n"
                        preview_text += "------------------------\n"
                        for mapping in mappings:
                            preview_text += f"Port: {mapping['port']}\n"
                            preview_text += f"Stow Code: {mapping['stow_code']}\n"
                            preview_text += "------------------------\n"
                
                self.preview_text.insert(tk.END, preview_text)
                
        except Exception as e:
            if hasattr(self, 'preview_text'):
                self.preview_text.delete(1.0, tk.END)
                self.preview_text.insert(tk.END, f"미리보기 업데이트 중 오류 발생: {str(e)}")

    def drop_tpsz_file(self, event):
        """TpSZ 엑셀 파일 드롭 처리"""
        try:
            file_path = event.data.strip('{}').strip('"')
            if not os.path.exists(file_path):
                messagebox.showerror("오류", "파일이 존재하지 않습니다.")
                return

            # 엑셀 파일 읽기
            df = pd.read_excel(file_path)
            
            # 컬럼 매핑 가져오기
            before_col = self.before_entry.get().strip()
            after_col = self.after_entry.get().strip()
            
            if not before_col or not after_col:
                messagebox.showerror("오류", "컬럼 매핑을 먼저 설정해주세요.")
                return
                
            if before_col not in df.columns or after_col not in df.columns:
                messagebox.showerror("오류", "설정한 컬럼명이 엑셀 파일에 존재하지 않습니다.")
                return

            # 매핑 딕셔너리 생성
            mapping = dict(zip(df[before_col], df[after_col]))
            
            # 설정 저장
            self.tpsz_mapping = mapping
            
            # JSON 파일 경로를 엑셀 파일과 동일한 디렉토리로 설정
            excel_dir = os.path.dirname(file_path)
            excel_name = os.path.splitext(os.path.basename(file_path))[0]
            self.tpsz_config_file = os.path.join(excel_dir, f"{excel_name}_mapping.json")
            
            self.save_tpsz_settings()
            
            # 미리보기 업데이트
            self.update_tpsz_preview()
            
            messagebox.showinfo("성공", "TpSZ 매핑이 성공적으로 업데이트되었습니다.")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")

    def save_tpsz_settings(self):
        """TpSZ 설정 저장"""
        try:
            # 현재 설정 가져오기
            settings = {
                'column_mapping': {
                    'before': self.before_entry.get().strip(),
                    'after': self.after_entry.get().strip()
                },
                'mapping': self.tpsz_mapping
            }
            
            # JSON 파일로 저장
            with open(self.tpsz_config_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
                
            messagebox.showinfo("성공", f"설정이 성공적으로 저장되었습니다.\n저장 위치: {self.tpsz_config_file}")
            
        except Exception as e:
            messagebox.showerror("오류", f"설정 저장 중 오류가 발생했습니다: {str(e)}")

    def update_tpsz_preview(self):
        """TpSZ 매핑 미리보기 업데이트"""
        try:
            if hasattr(self, 'tpsz_preview_text'):
                self.tpsz_preview_text.delete(1.0, tk.END)
                
                preview_text = "=== 컬럼 매핑 설정 ===\n"
                preview_text += f"Before 컬럼: {self.tpsz_column_mapping.get('before', '')}\n"
                preview_text += f"After 컬럼: {self.tpsz_column_mapping.get('after', '')}\n\n"
                
                preview_text += "=== 현재 매핑 ===\n"
                for before, after in self.tpsz_mapping.items():
                    preview_text += f"{before}: {after}\n"
                
                self.tpsz_preview_text.insert(tk.END, preview_text)
                
        except Exception as e:
            if hasattr(self, 'tpsz_preview_text'):
                self.tpsz_preview_text.delete(1.0, tk.END)
                self.tpsz_preview_text.insert(tk.END, f"미리보기 업데이트 중 오류 발생: {str(e)}")

    def setup_tpsz_tab(self):
        """TpSZ 관리 탭 설정"""
        # 메인 프레임
        main_frame = ttk.Frame(self.tpsz_tab)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 드래그 & 드롭 영역
        drop_frame = ttk.LabelFrame(main_frame, text="TpSZ 엑셀 파일")
        drop_frame.pack(fill="x", pady=(0, 10))

        self.tpsz_drop_label = ttk.Label(drop_frame, text="TpSZ 엑셀 파일을 여기에 드롭하세요")
        self.tpsz_drop_label.pack(pady=20)

        # 드래그 앤 드롭 바인딩
        self.tpsz_drop_label.drop_target_register(DND_FILES)
        self.tpsz_drop_label.dnd_bind('<<Drop>>', self.drop_tpsz_file)

        # 컬럼 매핑 설정 영역
        mapping_frame = ttk.LabelFrame(main_frame, text="컬럼 매핑 설정")
        mapping_frame.pack(fill="x", pady=(0, 10))

        # Before 컬럼 매핑
        before_frame = ttk.Frame(mapping_frame)
        before_frame.pack(fill="x", pady=5)
        ttk.Label(before_frame, text="Before 컬럼명:").pack(side="left", padx=5)
        self.before_entry = ttk.Entry(before_frame)
        self.before_entry.pack(side="left", fill="x", expand=True, padx=5)
        self.before_entry.insert(0, self.tpsz_column_mapping.get('before', ''))

        # After 컬럼 매핑
        after_frame = ttk.Frame(mapping_frame)
        after_frame.pack(fill="x", pady=5)
        ttk.Label(after_frame, text="After 컬럼명:").pack(side="left", padx=5)
        self.after_entry = ttk.Entry(after_frame)
        self.after_entry.pack(side="left", fill="x", expand=True, padx=5)
        self.after_entry.insert(0, self.tpsz_column_mapping.get('after', ''))

        # 저장 버튼
        save_button = ttk.Button(mapping_frame, text="설정 저장", command=self.save_tpsz_settings)
        save_button.pack(pady=10)

        # 현재 매핑 미리보기
        preview_frame = ttk.LabelFrame(main_frame, text="현재 매핑 미리보기")
        preview_frame.pack(fill="both", expand=True)
        
        self.tpsz_preview_text = tk.Text(preview_frame, height=10)
        self.tpsz_preview_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 컬럼 매핑 엔트리 값 설정
        if hasattr(self, 'before_entry'):
            self.before_entry.delete(0, tk.END)
            self.before_entry.insert(0, self.tpsz_column_mapping.get('before', ''))
        
        if hasattr(self, 'after_entry'):
            self.after_entry.delete(0, tk.END)
            self.after_entry.insert(0, self.tpsz_column_mapping.get('after', ''))
        
        # 미리보기 업데이트
        self.update_tpsz_preview()

    def terminal_to_port_mapping(self, terminal_code):
        # 터미널 코드에 따른 POL, TOL 매핑 딕셔너리
        terminal_mapping = {
            'PNITC': {'pol': 'KRPUS', 'tol': 'KRPUSAB'},
            'PNCOC': {'pol': 'KRPUS', 'tol': 'KRPUSPN'},
            'BCTHD': {'pol': 'KRPUS', 'tol': 'KRPUSBC'},
            'HJNPC': {'pol': 'KRPUS', 'tol': 'KRPUSAP'},
            'ICTPC': {'pol': 'KRINC', 'tol': 'KRINCAH'},
            'KEGWC': {'pol': 'KRKAN', 'tol': 'KRKANKT'}
        }
        
        return terminal_mapping.get(terminal_code, {'pol': '', 'tol': ''})

    def process_cll_file(self):
        # ... existing code ...
        
        # 엑셀 파일에서 (12,4) 위치의 터미널 코드 읽기
        terminal_code = worksheet.cell(12, 4).value
        
        # 터미널 코드를 기반으로 POL, TOL 값 설정
        port_info = self.terminal_to_port_mapping(terminal_code)
        self.pol_value = port_info['pol']
        self.tol_value = port_info['tol']
        
        # ... existing code ...

    def run(self):
        self.root.mainloop()

    def find_matching_services(self, pod_list):
        """POD 리스트와 일치하는 서비스 찾기"""
        matching_services = {}
        
        print(f"Finding matching services for POD list: {pod_list}")  # 디버깅용
        print(f"Available stow mappings: {self.stow_mapping}")  # 디버깅용
        
        # stow_mapping이 비어있는지 확인
        if not self.stow_mapping:
            print("Warning: stow_mapping is empty!")  # 디버깅용
            messagebox.showwarning("경고", "Stow Code 매핑 정보가 없습니다. Stowage Code 관리 탭에서 매핑 정보를 추가해주세요.")
            return matching_services
        
        for service_name, mappings in self.stow_mapping.items():
            print(f"Checking service: {service_name}")  # 디버깅용
            matches = []
            for pod in pod_list:
                print(f"  Checking POD: {pod}")  # 디버깅용
                for mapping in mappings:
                    print(f"    Comparing with mapping: {mapping}")  # 디버깅용
                    if pod.upper() == mapping['port'].upper():
                        print(f"      Found port match: {pod} = {mapping['port']}")  # 디버깅용
                        matches.append({
                            'pod': pod,
                            'port': mapping['port'],
                            'stow_code': mapping['stow_code']
                        })
                    elif pod.upper() == mapping['stow_code'].upper():
                        print(f"      Found stow_code match: {pod} = {mapping['stow_code']}")  # 디버깅용
                        matches.append({
                            'pod': pod,
                            'port': mapping['port'],
                            'stow_code': mapping['stow_code']
                        })
            if matches:
                print(f"  Found {len(matches)} matches for service {service_name}")  # 디버깅용
                matching_services[service_name] = matches
        
        print(f"Final matching services: {matching_services}")  # 디버깅용
        return matching_services

    def show_service_selection_dialog(self, matching_services):
        """서비스 선택 다이얼로그 표시"""
        dialog = tk.Toplevel(self.root)
        dialog.title("서비스 선택")
        dialog.geometry("600x400")
        
        # Stow Code 변환 미적용 버튼
        no_conversion_btn = tk.Button(
            dialog,
            text="Stow Code 변환 미적용",
            command=lambda: dialog.destroy(),
            relief="raised", 
            bg="#ff9999",  # 붉은 계열 배경색
            padx=10,
            pady=5,
            cursor="hand2"
        )
        no_conversion_btn.pack(pady=10)
        
        # 설명 레이블
        ttk.Label(dialog, text="발견된 POD와 일치하는 서비스 목록입니다.\n서비스를 클릭하여 선택해주세요.").pack(pady=10)
        
        # 스크롤 가능한 프레임 생성
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 캔버스 생성
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 결과 저장용 변수
        dialog.result = None
        no_conversion_btn.configure(command=lambda: [setattr(dialog, 'result', 'NONE SERVICE'), print("Selected service: NONE SERVICE"), dialog.destroy()])
        
        def on_service_click(service_name):
            dialog.result = service_name
            dialog.destroy()
        
        # 각 서비스에 대한 프레임 생성
        for service_name, matches in matching_services.items():
            # 서비스 프레임
            service_frame = ttk.Frame(scrollable_frame)
            service_frame.pack(fill="x", pady=5, padx=5)
            
            # 서비스 버튼 (클릭 시 바로 선택)
            btn = tk.Button(
                service_frame, 
                text=service_name,
                command=lambda sn=service_name: on_service_click(sn),
                relief="raised",
                bg="#e1e1e1",
                padx=10,
                pady=5,
                cursor="hand2"  # 마우스 오버 시 손가락 커서
            )
            btn.pack(fill="x")
            
            # 매칭 정보 표시
            info_text = tk.Text(service_frame, height=len(matches), wrap="word")
            info_text.pack(fill="x", padx=20)
            
            for match in matches:
                info_text.insert(tk.END, f"POD: {match['pod']} → Stow Code: {match['stow_code']}\n")
            
            info_text.config(state='disabled')  # 읽기 전용으로 설정
            
            # 구분선 추가
            ttk.Separator(scrollable_frame, orient='horizontal').pack(fill='x', pady=5)
        
        # 스크롤바와 캔버스 패킹
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        # 취소 버튼
        ttk.Button(dialog, text="취소", command=dialog.destroy).pack(pady=10)
        
        # 모달 대화상자로 실행
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.focus_set()
        
        # 창을 화면 중앙에 위치
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
        
        self.root.wait_window(dialog)
        result = dialog.result
        print(f"Selected service: {result}")  # 콘솔에 선택된 서비스 출력
        return result

    def apply_stow_codes(self, df, service_name):
        """선택된 서비스의 Stow Code 적용"""
        if not service_name or service_name not in self.stow_mapping:
            return df
        
        mappings = self.stow_mapping[service_name]
        
        # DataFrame 복사
        updated_df = df.copy()
        
        # 각 행에 대해 매핑 적용
        for idx, row in updated_df.iterrows():
            pod = str(row['POD']).strip().upper()
            fpod = str(row.get('FPOD', '')).strip().upper()
            stow = str(row.get('Stow', '')).strip().upper()

            
            # POD와 FPOD가 같으면 건너뛰기
            if pod != fpod:

                for mapping in mappings:
                    # if pod == mapping['port'].upper() or pod == mapping['stow_code'].upper():
                        updated_df.at[idx, 'POD'] = mapping['port']
                        updated_df.at[idx, 'Stow'] = mapping['stow_code']
                        break
        
        return updated_df

    def process_obl_file(self, file_path):
        """OBL 파일 처리"""
        try:
            # 엑셀 파일 읽기
            df = pd.read_excel(file_path)
            
            # POD 목록 추출
            pod_list = df['POD'].unique().tolist()
            
            # 매칭되는 서비스 찾기
            matching_services = self.find_matching_services(pod_list)
            
            if not matching_services:
                messagebox.showwarning("경고", "일치하는 서비스를 찾을 수 없습니다.")
                return
            
            # 서비스 선택 다이얼로그 표시
            selected_service = self.show_service_selection_dialog(matching_services)
            
            if not selected_service:
                return
            
            # Stow Code 적용
            updated_df = self.apply_stow_codes(df, selected_service)
            
            # 파일 저장
            # 엑셀 파일에서 vessel name과 항차 정보 읽기
            vessel_name = str(df.iloc[2, 2]).strip()  # C3 셀의 vessel name
            print(f"Vessel name: {vessel_name}")  # 디버깅용
            voyage_no = str(df.iloc[2, 11]).strip()   # L3 셀의 항차 정보
            print(f"Voyage no: {voyage_no}")  # 디버깅용
            # 파일명 생성
            save_path = os.path.join(
                os.path.dirname(file_path),
                f"{os.path.splitext(os.path.basename(file_path))[0]}_{vessel_name}_{voyage_no}_updated.xlsx"
            )
            updated_df.to_excel(save_path, index=False)
            
            messagebox.showinfo("성공", f"Stow Code가 적용되었습니다.\n저장 위치: {save_path}")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")

    def drop_obl_for_stow(self, event):
        """OBL 파일 드롭 처리"""
        file_path = event.data.strip('{}').strip('"')
        if not os.path.exists(file_path):
            messagebox.showerror("오류", "파일이 존재하지 않습니다.")
            return
        
        self.process_obl_file(file_path)

    def setup_edi_tab(self):
        """EDI PARSER 탭 설정"""
        # 좌우 분할을 위한 프레임
        left_frame = ttk.Frame(self.edi_tab)
        right_frame = ttk.Frame(self.edi_tab)
        left_frame.pack(side="left", fill="both", expand=True, padx=5)
        right_frame.pack(side="right", fill="both", expand=True, padx=5)

        # 왼쪽: EDI 파일 드래그 앤 드롭 영역
        drop_frame = ttk.LabelFrame(left_frame, text="EDI 파일 드래그 앤 드롭")
        drop_frame.pack(fill="both", expand=True, pady=5)

        self.edi_drop_label = ttk.Label(
            drop_frame,
            text="EDI 파일을 여기에 드롭하세요",
            font=('Arial', 12)
        )
        self.edi_drop_label.pack(fill="both", expand=True, padx=20, pady=20)

        # 드래그 앤 드롭 바인딩
        self.edi_drop_label.drop_target_register(DND_FILES)
        self.edi_drop_label.dnd_bind('<<Drop>>', self.process_edi_file)

        # 오른쪽: POD Summary 표시 영역
        summary_frame = ttk.LabelFrame(right_frame, text="POD 별 컨테이너 수량")
        summary_frame.pack(fill="both", expand=True, pady=5)

        self.pod_summary_text = tk.Text(
            summary_frame,
            height=20,
            width=40,
            font=('Courier', 10)
        )
        self.pod_summary_text.pack(fill="both", expand=True, padx=5, pady=5)

        # POD 표시를 위한 프레임 추가
        pod_frame = ttk.LabelFrame(self.edi_tab, text="추출된 POD", padding="5")
        pod_frame.pack(fill="x", padx=5, pady=5)
        
        # POD 값을 표시할 레이블 추가
        self.pod_label = ttk.Label(pod_frame, text="EDI 파일을 드래그 앤 드롭하세요")
        self.pod_label.pack(fill="x", padx=5, pady=5)

    def process_edi_file(self, event):
        try:
            # 클래스 변수 초기화
            self.OH = "0"
            self.Oleft = "0"
            self.Oright = "0"
            
            input_file_path = event.data.strip('{}')
            if not os.path.exists(input_file_path):
                messagebox.showerror("오류", "파일이 존재하지 않습니다.")
                return

            input_dir = os.path.dirname(input_file_path)
            
            # 엑셀 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"  # 시트 이름을 명확하게 지정

            # 기본 폰트 및 정렬 설정
            default_font = Font(name='Arial', size=10)
            default_alignment = Alignment(horizontal='left', vertical='center')
            
            # 기본 열 너비 설정
            for column in range(1, 26):  # A to Y
                ws.column_dimensions[utils.get_column_letter(column)].width = 12

            # 헤더 설정 (6번째 행)
            headers = ["POD", "CELL", "Cntr No.", "OPR", "POL", "STOW", "FPOD", "POR", 
                      "TpSz", "WGT", "F/E", "SP", "Temp", "DG", "UNNO", "PG", "FP",
                      "PrePos", "ACC.", "RSN", "Over Dimension", "Over Slot", "Remark",
                      "Void.Calc", "Void.Calc"]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=header)

            # 변수 초기화
            cntr_count = 6
            vessel = ""
            voy = ""
            port = ""
            formatted_date_time = ""

            # EDI 파일 읽기
            with open(input_file_path, 'r', encoding='utf-8') as f:
                file_content = f.read()
            
            # 라인별로 분리
            file_lines = file_content.replace('\r\n', '\n').split('\n')

            # EDI 라인 처리
            for line in file_lines:
                line = line.strip()
                if not line:
                    continue

                edi_lines = line.split('+')
                edi_lines2 = line.split('::')
                edi_lines3 = line.split(':')

                # DTM (날짜/시간) 처리
                if edi_lines[0] == "DTM":
                    date_time_parts = edi_lines[1].split(':')
                    if date_time_parts[0] == "137":
                        date_str = date_time_parts[1]
                        edi_date = f"{date_str[6:8]}.{date_str[4:6]}.{date_str[:4]}"
                        edi_time = f"{date_str[8:10]}:{date_str[10:12]}:23"
                        formatted_date_time = f"{edi_date} {edi_time}"
                        ws.cell(row=cntr_count-4, column=1, value=f"Vessel Name : {vessel}                                                                                                   Data : {formatted_date_time}")

                # TDT (선박 정보) 처리
                elif edi_lines[0] == "TDT":
                    vessel = edi_lines2[1][:-1]
                    voy = edi_lines[2]
                    
                    ws.cell(row=cntr_count-5, column=1, value="                                                               Inquary Summary(Detail Information)")
                    ws.cell(row=cntr_count-4, column=1, value=f"Vessel Name : {vessel}                                                                                                   Data : {formatted_date_time}")
                    ws.cell(row=cntr_count-2, column=1, value="Operator Code : ---")

                # LOC (위치 정보) 처리
                elif edi_lines[0] == "LOC":
                    # LOC+5 처리
                    if edi_lines[1] == "5":
                        port = edi_lines[2][:5]
                        ws.cell(row=cntr_count-3, column=1, value=f"Voyage No : {voy}                                                                                                   Port : {port}")

                    # 다른 LOC 타입 처리
                    loc_type = edi_lines[1]
                    cell_value = "UNSET"

                    if len(edi_lines) >= 3 and edi_lines[2].strip():
                        cell_value = int(edi_lines[2][:7]) if loc_type == "147" else edi_lines[2][:5]
                        if cell_value == "KRBUS":
                            cell_value = "KRPUS"

                        # LOC 타입별 처리
                        # 147: 컨테이너 번호 (2열)
                        if loc_type == "147":
                            cntr_count += 1  # 새로운 컨테이너 행 추가
                            ws.cell(row=cntr_count, column=2, value=cell_value)
                        # 9,6: 목적지 (5열) 
                        elif loc_type in ["9", "6"]:
                            ws.cell(row=cntr_count, column=5, value=cell_value)
                        # 11,12: 출발지 (1열)
                        elif loc_type in ["11", "12"]:
                            ws.cell(row=cntr_count, column=1, value=cell_value)
                        # 76: 최종 목적지 (8열)
                        elif loc_type == "76":
                            ws.cell(row=cntr_count, column=8, value=cell_value)
                        # 83: 환적항 (7열)
                        elif loc_type == "83":
                            ws.cell(row=cntr_count, column=7, value=cell_value)

                    # UNSET 값 설정
                    if not ws.cell(row=cntr_count, column=8).value:
                        ws.cell(row=cntr_count, column=8, value="UNSET")
                    if not ws.cell(row=cntr_count, column=7).value:
                        ws.cell(row=cntr_count, column=7, value="UNSET")
                    # F/E 값이 비어있는 경우 'F'로 설정
                    if not ws.cell(row=cntr_count, column=11).value:
                        ws.cell(row=cntr_count, column=11, value="F")

                # MEA (무게 정보) 처리
                elif edi_lines[0] == "MEA" and len(edi_lines) > 3:
                    try:
                        # MEA+WT++KGM:29600' 형식에서 무게 추출
                        weight_str = edi_lines[3].split(':')[1].rstrip("'")  # 29600 추출
                        weight = round(float(weight_str) / 1000, 1)  # 29600 → 29.6
                        ws.cell(row=cntr_count, column=10, value=weight)
                        ws.cell(row=cntr_count, column=10).number_format = "0.0"
                        # print(f"Weight processed: {weight}")  # 디버깅용
                    except Exception as e:
                        # print(f"Error processing weight: {str(e)}")  # 디버깅용
                        ws.cell(row=cntr_count, column=10, value="")

                # EQD (컨테이너 정보) 처리
                elif edi_lines[0] == "EQD":
                    try:
                        # 컨테이너 번호 저장 (3열)
                        ws.cell(row=cntr_count, column=3, value=edi_lines[2])
                        
                        # 컨테이너 타입 처리 (9열)
                        if len(edi_lines) > 3:
                            container_type = edi_lines[3]
                            type_mapping = {
                                # 20FT Containers
                                "2200": "20DV", "2210": "20DV", "22G0": "20DV", "22G1": "20DV",  # Standard 20ft
                                "22T0": "20TK", "22T1": "20TK",  # Tank 20ft
                                "2232": "20RE", "22R0": "20RE", "22R1": "20RE",  # Reefer 20ft
                                "22P1": "20FL", "22P0": "20FL",  # Flat Rack 20ft
                                "22U1": "20OT", "22U0": "20OT",  # Open Top 20ft
                                "22H0": "20HQ", "22H1": "20HQ",  # High Cube 20ft
                                "22B0": "20BK", "22B1": "20BK",  # Bulk 20ft
                                "2250": "20RF",  # Reefer 20ft
                                "22GP": "20GP", # General Purpose 20ft
                                "22PC": "20FR", # Platform Container 20ft
                                "22UT": "20OT", # Open Top 20ft

                                # 40FT Containers
                                "42G0": "40DV", "4310": "40DV", "42G1": "40DV",  # Standard 40ft
                                "45G0": "40HC", "4510": "40HC", "45G1": "40HC",  # High Cube 40ft
                                "45R0": "40HR", "4532": "40HR", "45R1": "40HR",  # High Cube Reefer 40ft
                                "42P1": "40FL", "4363": "40FL", "42P0": "40FL",  # Flat Rack 40ft
                                "42U1": "40OT", "42U0": "40OT",  # Open Top 40ft
                                "4232": "40RE", "42R0": "40RE",  # Reefer 40ft
                                "42T0": "40TK", "42T1": "40TK",  # Tank 40ft
                                "4563": "40HF",  # High Cube Flat Rack 40ft
                                "42B0": "40BK",  # Bulk 40ft
                                "40GP": "40GP", # General Purpose 40ft
                                "40PC": "40FR", # Platform Container 40ft
                                "40UT": "40OT", # Open Top 40ft
                                "43GP": "40HC", # 40ft High Cube

                                # 45FT Containers
                                "9400": "45HC", "L5G0": "45HC",  # High Cube 45ft
                                "L5G1": "45HC", "95G0": "45HC",  # High Cube 45ft variants
                                "45GP": "45HC", # 45ft High Cube General Purpose
                                "45PC": "45FR", # Platform Container 45ft
                                "45UT": "45OT", # Open Top 45ft

                                # Special Equipment
                                "GENE": "GE",  # Generator
                                "VENT": "VT",  # Ventilated
                                "CONT": "CT",  # Controlled Temperature
                                "CRYO": "CY",  # Cryogenic
                                "HCFR": "HRF", # High Cube Flat Rack
                                "PCHP": "HP", # Platform
                                "REOT": "RO", # Reefer Open Top
                                "TKOT": "TO", # Tank Open Top
                                "PCOT": "PO", # Platform Open Top
                                "FLOT": "FO", # Flat Rack Open Top
                                "SKEL": "SK", # Skeletal
                                "FRMG": "FG", # Frame
                                "BULD": "BD", # Bulked
                                "LIVS": "LS", # Live Stock
                                "VEHI": "VH", # Vehicle Carrier
                                "PIPE": "PP", # Pipe Carrier
                                "LOGS": "LG", # Log Carrier
                                "DANG": "DG", # Dangerous Goods
                                "EXPL": "EX", # Explosives
                                "RADIO": "RD", # Radioactive
                                "OXID": "OX", # Oxidizing Substances
                                "CORR": "CR", # Corrosives
                                "MISC": "MC", # Miscellaneous Dangerous Goods
                                "20HC": "20HC", # 20ft High Cube
                                "40PW": "40PW", # 40ft Pallet Wide
                                "45PW": "45PW", # 45ft Pallet Wide
                                "20RF": "20RF", # 20ft Reefer
                                "40RF": "40RF", # 40ft Reefer
                                "45RF": "45RF", # 45ft Reefer
                                "20TN": "20TN", # 20ft Tank
                                "40TN": "40TN", # 40ft Tank
                                "20PL": "20PL", # 20ft Platform
                                "40PL": "40PL", # 40ft Platform
                                "45PL": "45PL", # 45ft Platform
                                "20OS": "20OS", # 20ft Open Side
                                "40OS": "40OS", # 40ft Open Side
                                "20VN": "20VN", # 20ft Ventilated
                                "40VN": "40VN", # 40ft Ventilatedㅇ뎀ㄱ셕ㄷ
                                "20SS": "20SS", # 20ft Side Stanchion
                                "40SS": "40SS", # 40ft Side Stanchion
                                "20HT": "20HT", # 20ft Hard Top
                                "40HT": "40HT", # 40ft Hard Top
                                "20OT": "20OT", # 20ft Open Top
                                "40OT": "40OT", # 40ft Open Top
                                "40HF": "40HF", # 40ft Open Top
                                "40HO": "40HO", # 40ft Open Top
                                "45OT": "45OT", # 45ft Open Top
                            }
                            new_type = type_mapping.get(container_type, "")
                            
                            # If container_type not in mapping, check if it starts with 2 or 4
                            if not new_type and container_type:
                                if container_type.startswith('2'):
                                    new_type = "20DV"
                                elif container_type.startswith('4'):
                                    new_type = "40HC"
                                    
                            ws.cell(row=cntr_count, column=9, value=new_type)

                        # E/F 상태 처리 (11열)
                        if len(edi_lines) > 6:
                            last_value = edi_lines[6].rstrip("'")  # 마지막 따옴표 제거
                            if last_value == "4":
                                ws.cell(row=cntr_count, column=11, value="E")
                            elif last_value == "5":
                                ws.cell(row=cntr_count, column=11, value="F")
                        
                        # print(f"EQD processed - Container: {edi_lines[2]}, Type: {new_type}, Status: {ws.cell(row=cntr_count, column=11).value}")  # 디버깅용
                    except Exception as e:
                        print(f"Error processing EQD: {str(e)}")  # 디버깅용

                # CN (컨테이너 상태) 처리
                elif edi_lines[0] == "CN" and len(edi_lines) > 6:
                    if edi_lines[6] == "4'":
                        ws.cell(row=cntr_count, column=11, value="E")
                    elif edi_lines[6] == "5'":
                        ws.cell(row=cntr_count, column=11, value="F")

                # NAD (운송인 정보) 처리
                elif edi_lines[0] == "NAD" and len(edi_lines) > 2:
                    ws.cell(row=cntr_count, column=4, value=edi_lines[2][:3])

                # TMP (온도 정보) 처리
                elif edi_lines[0] == "TMP" and len(edi_lines) > 2:
                    try:
                        # TMP+2+05.0:CEL' 또는 TMP+2+00.0:CEL' 형식에서 온도값 추출
                        temp_str = edi_lines[2].split(':')[0].strip()  # :CEL' 부분 제거
                        
                        # 부호 처리
                        is_negative = temp_str.startswith('-')
                        if is_negative:
                            temp_str = temp_str[1:]  # 마이너스 부호 제거
                        elif temp_str.startswith('+'):
                            temp_str = temp_str[1:]  # 플러스 부호 제거
                            
                        # 앞의 0 제거하고 소수점 처리
                        if temp_str.startswith('0') and not temp_str.startswith('0.'):
                            temp_str = temp_str[1:]  # 앞의 0 제거
                            
                        # 부호 다시 추가
                        if is_negative:
                            temp_str = f"-{temp_str}"
                            
                        ws.cell(row=cntr_count, column=13, value=f"{temp_str}C")
                        # print(f"Temperature processed: {temp_str}C")  # 디버깅용
                    except Exception as e:
                        print(f"Error processing temperature: {str(e)}")  # 디버깅용
                        ws.cell(row=cntr_count, column=13, value="")  # 에러 시 빈 값 설정

                # DGS (위험물 정보) 처리
                elif edi_lines[0] == "DGS" and len(edi_lines) > 3:
                    ws.cell(row=cntr_count, column=14, value=float(edi_lines[2]))
                    ws.cell(row=cntr_count, column=15, value=float(edi_lines[3]))
                # DIM (치수 정보) 처리
                elif edi_lines[0] == "DIM" and len(edi_lines) > 2:
                    # 전역 변수로 OH, Oleft, Oright 선언
                    global OH, Oleft, Oright
                    
                    if edi_lines[1] == "9":
                        self.OH = str(int(edi_lines[2].split(":::")[1].rstrip("'")))  # Convert to int to remove leading zeros, then back to string
                    elif edi_lines[1] == "8":
                        self.Oleft = str(int(edi_lines[2].split("::")[1].rstrip("'")))  # Convert to int to remove leading zeros, then back to string
                    elif edi_lines[1] == "7":
                        self.Oright = str(int(edi_lines[2].split("::")[1].rstrip("'")))  # Convert to int to remove leading zeros, then back to string
                    
                    print(f"OH: {self.OH}, Oleft: {self.Oleft}, Oright: {self.Oright}")
                    ws.cell(row=cntr_count, column=21, value=f"//{self.OH}/{self.Oleft}/{self.Oright}")
          

            # POD 요약 생성
            pod_summary = {}
            pol_pod_summary = {}  # POL과 port가 일치하는 데이터의 요약
            for row in range(7, ws.max_row + 1):  # 데이터가 시작되는 7행부터
                pod = ws.cell(row=row, column=1).value  # POD는 1열(A열)에 있음
                pol = ws.cell(row=row, column=5).value  # POL은 5열(E열)에 있음
                
                if pod and pod != "UNSET":  # POD 값이 있고 UNSET이 아닌 경우만
                    pod_summary[pod] = pod_summary.get(pod, 0) + 1
                    # POL이 현재 port와 일치하는 경우만 별도 집계
                    if pol == port:
                        pol_pod_summary[pod] = pol_pod_summary.get(pod, 0) + 1

            # POD 요약 텍스트 업데이트
            self.pod_summary_text.delete(1.0, tk.END)
            
            # 색상 태그 설정 (배경색과 보색)
            self.pod_summary_text.tag_configure("krpus", 
                background="#90EE90",  # 연한 녹색 배경
                foreground="#FF1493")  # 진한 분홍색 글자
            
            self.pod_summary_text.tag_configure("krkan", 
                background="#FFD700",  # 골드 배경
                foreground="#000080")  # 네이비 글자
            
            self.pod_summary_text.tag_configure("krinc", 
                background="#87CEEB",  # 하늘색 배경
                foreground="#FF4500")  # 주황색 글자
            
            # 선박 및 항차 정보 추가
            self.pod_summary_text.insert(tk.END, f"Vessel:  {vessel}\n")
            self.pod_summary_text.insert(tk.END, f"Voyage:  {voy}\n") 
            self.pod_summary_text.insert(tk.END, f"Port:    {port}\n\n")
            
            # 전체 POD Summary 출력
            self.pod_summary_text.insert(tk.END, "=== Total POD Summary ===\n\n")
            
            total_containers = 0
            for pod, count in sorted(pod_summary.items()):  # POD 알파벳 순으로 정렬
                # POD별로 다른 배경색과 글자색 적용
                if pod == "KRPUS":
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n", "krpus")
                elif pod == "KRKAN":
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n", "krkan")
                elif pod == "KRINC":
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n", "krinc")
                else:
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n")
                total_containers += count
            
            self.pod_summary_text.insert(tk.END, f"\nTotal: {total_containers}\n\n")

            # POL 기준 POD Summary 출력
            self.pod_summary_text.insert(tk.END, f"=== POD Summary (From {port}) ===\n\n")
            
            pol_total_containers = 0
            for pod, count in sorted(pol_pod_summary.items()):  # POD 알파벳 순으로 정렬
                # POD별로 다른 배경색과 글자색 적용
                if pod == "KRPUS":
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n", "krpus")
                elif pod == "KRKAN":
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n", "krkan")
                elif pod == "KRINC":
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n", "krinc")
                else:
                    self.pod_summary_text.insert(tk.END, f"{pod}: {count}\n")
                pol_total_containers += count
            
            self.pod_summary_text.insert(tk.END, f"\nTotal from {port}: {pol_total_containers}")

            # 파일 저장
            output_filename = f"{vessel} {voy} {port}.xlsx"
            output_file = os.path.join(input_dir, output_filename)
            
            # 기존 파일이 있다면 삭제
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                except PermissionError:
                    messagebox.showerror("오류", "기존 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요.")
                    return

            try:
                wb.save(output_file)
                print(f"File saved successfully at: {output_file}")  # 디버깅용
                
                # 파일이 정상적으로 생성되었는지 확인
                if os.path.exists(output_file):
                    file_size = os.path.getsize(output_file)
                    print(f"Created file size: {file_size} bytes")  # 디버깅용
                    if file_size > 0:
                        messagebox.showinfo("성공", f"EDI 파일이 성공적으로 변환되었습니다.\n저장 위치: {output_file}")
                    else:
                        messagebox.showerror("오류", "파일이 올바르게 생성되지 않았습니다.")
                else:
                    messagebox.showerror("오류", "파일 생성에 실패했습니다.")
            except Exception as e:
                print(f"Error saving file: {str(e)}")  # 디버깅용
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {str(e)}")
            finally:
                wb.close()

            # EDI 파일 처리 로직
            with open(input_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                
            # POD 추출 - 모든 매칭되는 값 찾기
            pod_values = set()  # 중복 제거를 위해 set 사용
            for line in content.split('\n'):
                if '+11+' in line:
                    parts = line.split('+')
                    if len(parts) >= 3:
                        pod = parts[2][:5]  # 5자리만 추출
                        pod_values.add(pod)
                        print(f"Extracted POD: {pod}")  # 디버깅용
            
            # POD 값들을 정렬하여 표시
            sorted_pods = sorted(list(pod_values))
            pod_text = "\n".join(sorted_pods)
            self.pod_label.config(text=pod_text)
            
            if pod_values:
                print(f"Found POD values: {pod_values}")  # 디버깅용
                
                # 매칭되는 서비스 찾기
                print(f"Calling find_matching_services with pod_values: {pod_values}")  # 디버깅용
                matching_services = self.find_matching_services(pod_values)
                
                # 서비스 선택 다이얼로그 표시 (매칭되는 서비스가 없더라도 표시)
                print("Calling show_service_selection_dialog")  # 디버깅용
                
                # 매칭되는 서비스가 없는 경우 빈 딕셔너리 대신 기본 서비스 추가
                if not matching_services:
                    print("No matching services found, adding default service")  # 디버깅용
                    matching_services = {"기본 서비스": [{"pod": pod, "port": pod, "stow_code": pod} for pod in pod_values]}
                
                selected_service = self.show_service_selection_dialog(matching_services)
                
                if not selected_service:
                    print("No service selected")  # 디버깅용
                    return
                
                print(f"Selected service: {selected_service}")  # 디버깅용
                
                # 선택된 서비스의 매핑 가져오기
                service_mappings = self.stow_mapping.get(selected_service, [])
                print(f"Service mappings: {service_mappings}")  # 디버깅용
                
                # 엑셀 파일 다시 열기
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active
                
                # 각 POD에 대해 매핑 적용
                for row in range(7, ws.max_row + 1):
                    pod = ws.cell(row=row, column=1).value  # POD는 1열(A열)에 있음
                    if pod and pod != "UNSET":
                        print(f"Processing POD: {pod}")  # 디버깅용
                        # POD에 대한 매핑 확인
                        for mapping in service_mappings:
                            print(f"Checking mapping: {mapping}")  # 디버깅용
                            if pod.upper() == mapping['stow_code'].upper():
                                print(f"Found stow_code match: {pod} = {mapping['stow_code']}")  # 디버깅용
                                # stow_code가 일치하면 해당 port를 POD로 사용하고 stow_code를 Stow 값으로 설정
                                ws.cell(row=row, column=1, value=mapping['port'])  # POD 열 업데이트
                                ws.cell(row=row, column=6, value=mapping['stow_code'])  # Stow 열 업데이트
                                break
                            elif pod.upper() == mapping['port'].upper():
                                print(f"Found port match: {pod} = {mapping['port']}")  # 디버깅용
                                # port가 일치하면 해당 stow_code 사용
                                ws.cell(row=row, column=1, value=mapping['port'])  # POD 열 업데이트
                                ws.cell(row=row, column=6, value=mapping['stow_code'])  # Stow 열 업데이트
                                break
                
                # 파일 저장
                wb.save(output_file)
                wb.close()
                
                messagebox.showinfo("성공", f"Stow Code가 적용되었습니다.")
                
                pod_text = "\n".join(sorted(pod_values))  # 정렬된 고유값들을 세로로 나열
                
                # POD 클릭 이벤트 핸들러 
                def on_pod_click(event):
                    # Label의 현재 텍스트 가져오기
                    clicked_text = event.widget.cget("text")
                    # 클릭된 위치의 y 좌표를 기반으로 라인 계산
                    y_position = event.y
                    line_height = event.widget.winfo_height() / len(clicked_text.split('\n'))
                    clicked_line_index = int(y_position / line_height)
                    # 클릭된 라인의 텍스트 가져오기
                    clicked_line = clicked_text.split('\n')[clicked_line_index]
                    
                    if clicked_line:
                        selected_pod = clicked_line.strip()
                        # 새로운 POD 값 입력 다이얼로그
                        new_pod = tk.simpledialog.askstring("POD 변경", 
                            f"현재 POD: {selected_pod}\n새로운 POD 값을 입력하세요:",
                            parent=self.root)
                        
                        if new_pod:
                            # 엑셀 파일 다시 열기
                            wb = openpyxl.load_workbook(output_file)
                            ws = wb.active
                            
                            # POD 값 변경
                            for row in range(7, ws.max_row + 1):
                                if ws.cell(row=row, column=1).value == selected_pod:
                                    ws.cell(row=row, column=6).value = selected_pod
                                    ws.cell(row=row, column=1).value = new_pod
                            
                            # 파일 저장
                            wb.save(output_file)
                            wb.close()
                            
                            # POD 라벨 업데이트
                            pod_values.remove(selected_pod)
                            pod_values.add(new_pod)
                            new_pod_text = "\n".join(sorted(pod_values))
                            self.pod_label.config(text=f"{new_pod_text}\n")
                            
                            messagebox.showinfo("성공", f"POD가 {selected_pod}에서 {new_pod}로 변경되었습니다.")
                
                # 클릭 이벤트 바인딩
                self.pod_label.bind('<Button-1>', on_pod_click)

                self.pod_label.config(text=f"{pod_text}\n")
            else:
                self.pod_label.config(text="POD를 찾을 수 없습니다")
        except Exception as e:
            print(f"Error processing EDI file: {str(e)}")
            messagebox.showerror("오류", f"EDI 파일 처리 중 오류가 발생했습니다: {str(e)}")

if __name__ == "__main__":
    app = ContainerConverter()
    app.run()


#test5